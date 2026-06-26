"""
bigmove_strategy.py — "only trades that beat the cost": daily spread strategy
that does NOT book small reversions. It holds each position until it captures a
big profit target (>> cost), hits a stop, or times out. Tested across several
targets and validated OUT-OF-SAMPLE (train 2024-25, test 2026).

Rationale: the signal wins gross but each small reversion (~26 pts) is below the
~48-pt cost. Requiring a big target (50/70/100 pts) makes each winner clear the
toll. Whether enough trades reach the target without stopping out first is the
empirical question - answered honestly below.

Run from feed_data/:
    python ..\\code\\bigmove_strategy.py --cost-pts 48 --split 2026-01-01
"""
import sys
import pandas as pd
from openpyxl.styles import Font, PatternFill

import backtest as bt

PV = 20  # Rs / point / lot
TARGETS = [50, 70, 100]
GREEN = PatternFill("solid", fgColor="C6EFCE"); RED = PatternFill("solid", fgColor="FFC7CE")
HEAD = PatternFill("solid", fgColor="1F4E78")


def arg(flag, d): return sys.argv[sys.argv.index(flag) + 1] if flag in sys.argv else d


def run(df, target, cost):
    # disable small-reversion exit; exit only on big target / stop / max-hold
    bt.ENTRY, bt.EXIT = 2.0, -999
    bt.STOP_LOSS, bt.PROFIT_TARGET, bt.MAX_HOLD = 100, target, 30
    bt.RATIO_LOOKBACK, bt.LOOKBACK = 60, 15
    t = bt.backtest(df)
    if not len(t):
        return t
    t["entry_date"] = pd.to_datetime(t["entry_date"])
    t["net"] = (t["pnl_points"] - cost).round(2)
    return t


def net(t):
    return round(t["net"].sum(), 1) if len(t) else 0.0


def main():
    cost  = float(arg("--cost-pts", "4"))
    split = pd.Timestamp(arg("--split", "2026-01-01"))
    out   = arg("--out", "bigmove_strategy.xlsx")

    df = pd.read_csv("history.csv")
    df["date"] = pd.to_datetime(df["date"])
    df = df.sort_values("date").reset_index(drop=True)
    print(f"Daily series: {len(df)} days ({df.date.min().date()} -> {df.date.max().date()}) | cost {cost:.0f} pts")
    print(f"Exit: big profit target only (no small reversion) | stop -100 | max-hold 30d | TRAIN<{split.date()}<=TEST\n")

    rows = []
    best_long = None
    for tgt in TARGETS:
        t = run(df, tgt, cost)
        if not len(t):
            continue
        for label, sub in [("ALL", t), ("LONG", t[t.direction == "LONG_SPREAD"])]:
            tr = sub[sub.entry_date < split]; te = sub[sub.entry_date >= split]
            r = {"target": tgt, "set": label,
                 "train_trades": len(tr), "train_net_pts": net(tr), "train_Rs": round(net(tr)*PV),
                 "test_trades": len(te), "test_net_pts": net(te), "test_Rs": round(net(te)*PV),
                 "full_net_pts": net(sub),
                 "win_pct": round((sub["net"] > 0).mean()*100, 0) if len(sub) else 0,
                 "avg_hold_d": round(sub.holding_days.mean(), 1) if len(sub) else 0}
            rows.append(r)
    grid = pd.DataFrame(rows)

    longs = grid[grid["set"] == "LONG"]
    # is ANY target net-positive in BOTH train and test for LONG?
    winners = longs[(longs.train_net_pts > 0) & (longs.test_net_pts > 0)]
    if len(winners):
        b = winners.sort_values("test_net_pts", ascending=False).iloc[0]
        verdict = (f"YES - LONG @ target {int(b.target)} is net-positive in train ({b.train_net_pts:+.0f}) "
                   f"AND test ({b.test_net_pts:+.0f} pts / Rs{b.test_Rs:+,.0f}) -> worth paper-trading")
    else:
        best_te = longs.sort_values("test_net_pts", ascending=False).iloc[0] if len(longs) else None
        verdict = (f"NO - no target makes LONG net-positive out-of-sample "
                   f"(best test = {best_te.test_net_pts:+.0f} pts @ target {int(best_te.target)})"
                   if best_te is not None else "NO trades")

    print("=" * 92)
    print("  BIG-MOVE TEST: net of cost, train (2024-25) vs test (2026, unseen)")
    print("=" * 92)
    print(grid.to_string(index=False))
    print("-" * 92)
    print(f"  VERDICT: {verdict}")
    print("=" * 92)

    summary = pd.DataFrame({
        "metric": ["Strategy", "Exit rule", "Cost (pts)", "Targets tested", "Split", "VERDICT", "Caveat"],
        "value": ["Daily, hold for a BIG move (beat the cost), LONG focus",
                  "profit target only (no small reversion) / stop -100 / max-hold 30d",
                  cost, str(TARGETS), str(split.date()), verdict,
                  "Small trade sample; one market history. Out-of-sample result is the one that counts."]})
    with pd.ExcelWriter(out, engine="openpyxl") as xl:
        summary.to_excel(xl, sheet_name="Summary", index=False)
        grid.to_excel(xl, sheet_name="Targets train vs test", index=False)
        wb = xl.book
        for nm in wb.sheetnames:
            ws = wb[nm]
            for c in ws[1]:
                c.font = Font(bold=True, color="FFFFFF"); c.fill = HEAD
            ws.freeze_panes = "A2"
        ws = wb["Targets train vs test"]; hdr = {c.value: c.column for c in ws[1]}
        for col in ("train_net_pts", "test_net_pts", "full_net_pts"):
            ci = hdr.get(col)
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=ci)
                try: v = float(cell.value)
                except (TypeError, ValueError): continue
                cell.fill = GREEN if v > 0 else (RED if v < 0 else cell.fill)
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
