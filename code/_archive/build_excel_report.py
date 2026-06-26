"""
build_excel_report.py — consolidate today's backtest into ONE Excel workbook
with a P&L summary plus all trade and configuration details.

Sheets:
  Summary               - plain-language verdict + key P&L numbers
  Trades (1min)         - the realistic 11-trade log with gross/cost/NET P&L
  Trades (per-second)   - the raw per-second trade log (no costs)
  All Configs           - full parameter sweep leaderboard (gross & net)
  Profitable Configs    - net-profitable settings (flagged as single-trade flukes)
  Price Data (1min)     - SENSEX/NIFTY 1-minute closes used by the backtest

Output: backtest_2026_06_18_report.xlsx
"""
import pandas as pd
import sys
from datetime import date
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from backtest_intraday import load_series, resample


def arg(flag, default):
    return sys.argv[sys.argv.index(flag) + 1] if flag in sys.argv else default


GREEN = PatternFill("solid", fgColor="C6EFCE")
RED   = PatternFill("solid", fgColor="FFC7CE")
HEAD  = PatternFill("solid", fgColor="1F4E78")
BANNER = PatternFill("solid", fgColor="FFF2CC")


def read(path):
    try:
        return pd.read_csv(path)
    except Exception:
        return None


def autoformat(ws, money_cols=(), result_col=None):
    """Bold header, freeze top row, size columns, colour P&L cells."""
    # header style
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEAD
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"

    # column widths
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(width + 2, 10), 46)

    # header -> index map
    headers = {c.value: c.column for c in ws[1]}
    for name in money_cols:
        ci = headers.get(name)
        if not ci:
            continue
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=ci)
            try:
                v = float(cell.value)
            except (TypeError, ValueError):
                continue
            if v > 0:
                cell.fill = GREEN
            elif v < 0:
                cell.fill = RED
            cell.number_format = "+#,##0.00;-#,##0.00"
    if result_col:
        ci = headers.get(result_col)
        if ci:
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=ci)
                if str(cell.value).upper() == "WIN":
                    cell.fill = GREEN
                elif str(cell.value).upper() == "LOSS":
                    cell.fill = RED


def main():
    tag = arg("--tag", date.today().strftime("%Y_%m_%d"))
    OUT = f"backtest_{tag}_report.xlsx"
    summary  = read(f"backtest_{tag}_SUMMARY.csv")
    t1min    = read(f"backtest_{tag}_1min.csv")
    tpersec  = read(f"backtest_{tag}.csv")
    allcfg   = read(f"backtest_{tag}_all_configs.csv")
    profit   = read(f"backtest_{tag}_configs.csv")

    # 1-min price context
    px = resample(load_series(arg("--sensex", "sensex_data.csv"), arg("--nifty", "nifty_data.csv")), "1min")
    px = px.rename(columns={"date": "time", "sensex_close": "sensex", "nifty_close": "nifty"})

    with pd.ExcelWriter(OUT, engine="openpyxl") as xl:
        if summary is not None: summary.to_excel(xl, sheet_name="Summary", index=False)
        if t1min   is not None: t1min.to_excel(xl,   sheet_name="Trades (1min)", index=False)
        if tpersec is not None: tpersec.to_excel(xl, sheet_name="Trades (per-second)", index=False)
        if allcfg  is not None: allcfg.to_excel(xl,  sheet_name="All Configs", index=False)
        if profit  is not None: profit.to_excel(xl,  sheet_name="Profitable Configs", index=False)
        px.to_excel(xl, sheet_name="Price Data (1min)", index=False)

        wb = xl.book

        # Summary sheet — special banner formatting
        ws = wb["Summary"]
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF"); cell.fill = HEAD
        ws.freeze_panes = "A2"
        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 26
        ws.column_dimensions["C"].width = 80
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=1).font = Font(bold=True)
            ws.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical="top")
            label = str(ws.cell(row=r, column=1).value or "")
            if label in ("Net P&L", "VERDICT"):
                for c in range(1, 4):
                    ws.cell(row=r, column=c).fill = RED
            elif label == "Gross P&L":
                for c in range(1, 4):
                    ws.cell(row=r, column=c).fill = BANNER

        # Trade sheets — colour P&L + result
        if t1min is not None:
            autoformat(wb["Trades (1min)"],
                       money_cols=("gross_pnl_points", "net_pnl_points", "running_net_pts"),
                       result_col="result")
        if tpersec is not None:
            pcols = tuple(c for c in ("pnl_points", "net_pnl_points", "cum_pnl_points", "cum_net_points")
                          if tpersec is not None and c in tpersec.columns)
            autoformat(wb["Trades (per-second)"], money_cols=pcols, result_col="result")
        autoformat(wb["All Configs"], money_cols=("gross_pts", "net_pts"))
        autoformat(wb["Profitable Configs"], money_cols=("gross_pts", "net_pts"))
        autoformat(wb["Price Data (1min)"])

    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
