"""
feed_to_excel.py — dump the live per-second feed CSVs into ONE Excel workbook
with every field captured (run at end of day or any time).

The live feed writes ticks to CSV (append-safe). Excel cannot be appended to
one row per second, so we export the full CSVs to .xlsx here.

Usage:
    python feed_to_excel.py                  # today's sensex_data.csv / nifty_data.csv
    python feed_to_excel.py --tag 2026_06_22 --out feed_data_2026_06_22.xlsx
"""
import sys
from datetime import date
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEAD = PatternFill("solid", fgColor="1F4E78")


def arg(flag, default):
    return sys.argv[sys.argv.index(flag) + 1] if flag in sys.argv else default


def fmt(ws):
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEAD
        c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    for col in ws.columns:
        name = str(col[0].value)
        width = 46 if name == "raw_json" else min(
            max((len(str(c.value)) for c in col[:200] if c.value is not None), default=10) + 2, 30)
        ws.column_dimensions[get_column_letter(col[0].column)].width = width


def main():
    tag = arg("--tag", date.today().strftime("%Y_%m_%d"))
    out = arg("--out", f"feed_data_{tag}.xlsx")
    sx = pd.read_csv(arg("--sensex", "sensex_data.csv"))
    nf = pd.read_csv(arg("--nifty",  "nifty_data.csv"))

    with pd.ExcelWriter(out, engine="openpyxl") as xl:
        sx.to_excel(xl, sheet_name="SENSEX", index=False)
        nf.to_excel(xl, sheet_name="NIFTY", index=False)
        fmt(xl.book["SENSEX"])
        fmt(xl.book["NIFTY"])

    print(f"Wrote {out}  | SENSEX {len(sx)} ticks, NIFTY {len(nf)} ticks, {len(sx.columns)} fields each")


if __name__ == "__main__":
    main()
