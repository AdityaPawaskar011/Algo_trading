"""
backtest_daily.py — proper MULTI-DAY (position-holding) backtest of the spread
strategy on DAILY data, with transaction costs.

This is the timeframe the strategy was designed for: enter at |z| >= ENTRY and
HOLD the position across days (futures are carried overnight, squared off only
near monthly expiry) until the spread reverts (|z| <= EXIT), hits a stop/target,
or MAX_HOLD days pass. Each trade captures a large multi-day move, so the fixed
~35-pt round-trip cost is small relative to the profit.

Data: history.csv (daily SENSEX/NIFTY closes) + the recent live-feed days' closes.

Usage:
    python backtest_daily.py --cost-pts 35
    python backtest_daily.py --cost-pts 35 --target 200 --maxhold 25   # let winners run
"""
import sys
import pandas as pd
from openpyxl.styles import Font, PatternFill

import backtest as bt

FEED_DAYS = [
    ("2026-06-18", "sensex_Nifty/sensex_data.csv",        "sensex_Nifty/nifty_data.csv"),
    ("2026-06-19", "sensex_data_2026-06-19_archived.csv", "nifty_data_2026-06-19_archived.csv"),
    ("2026-06-22", "sensex_data.csv",                     "nifty_data.csv"),
    ("2026-06-23", "sensex_today.csv",                    "nifty_today.csv"),
]


def arg(flag, default):
    return sys.argv[sys.argv.index(flag) + 1] if flag in sys.argv else default


def daily_close(path):
    df = pd.read_csv(path, usecols=["tick_time", "last_price"]).dropna()
    return float(df.iloc[-1]["last_price"])


def build_daily():
    """history.csv (Yahoo daily) + recent feed days' closing prices."""
    h = pd.read_csv("history.csv")
    h["date"] = pd.to_datetime(h["date"])
    extra = []
    for d, sxf, nff in FEED_DAYS:
        try:
            extra.append({"date": pd.Timestamp(d),
                          "sensex_close": daily_close(sxf),
                          "nifty_close": daily_close(nff)})
        except Exception:
            pass
    if extra:
        h = pd.concat([h, pd.DataFrame(extra)], ignore_index=True)
    h = h.drop_duplicates("date").sort_values("date").reset_index(drop=True)
    return h[["date", "sensex_close", "nifty_close"]]


def fmt(ws):
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E78")
    ws.freeze_panes = "A2"


def main():
    cost = float(arg("--cost-pts", "4"))
    # optional overrides so winners can run further than the default +30 target
    if "--target" in sys.argv:
        bt.PROFIT_TARGET = float(arg("--target", "30"))
    if "--maxhold" in sys.argv:
        bt.MAX_HOLD = int(arg("--maxhold", "15"))

    df = build_daily()
    print(f"Daily series: {len(df)} days ({df.date.min().date()} -> {df.date.max().date()})")
    print(f"Strategy: entry +/-{bt.ENTRY} exit +/-{bt.EXIT} | SL -{bt.STOP_LOSS} "
          f"TP +{bt.PROFIT_TARGET} | max-hold {bt.MAX_HOLD} days | cost {cost:.0f} pts/round-trip")

    tr = bt.backtest(df)
    if not len(tr):
        print("No trades generated.")
        return

    tr = tr.rename(columns={"pnl_points": "gross_pnl_points"})
    tr["cost_points"]    = cost
    tr["net_pnl_points"] = (tr["gross_pnl_points"] - cost).round(2)
    tr["cum_net_points"] = tr["net_pnl_points"].cumsum().round(2)
    tr["result"] = tr["net_pnl_points"].apply(lambda p: "WIN" if p > 0 else "LOSS")
    cols = ["entry_date", "exit_date", "direction", "spread_type", "holding_days",
            "sensex_entry", "nifty_entry", "ratio", "spread_entry", "zscore_entry",
            "spread_exit", "zscore_exit", "exit_reason",
            "gross_pnl_points", "cost_points", "net_pnl_points", "cum_net_points", "result"]
    tr = tr[[c for c in cols if c in tr.columns]]
    tr.to_csv("backtest_daily_trades.csv", index=False)

    g, n = tr.gross_pnl_points, tr.net_pnl_points
    print("\n" + "=" * 70)
    print("  MULTI-DAY (position-holding) SPREAD BACKTEST")
    print("=" * 70)
    print(f"  Trades            : {len(tr)}")
    print(f"  Avg / max hold    : {tr.holding_days.mean():.1f} / {tr.holding_days.max()} days")
    print(f"  GROSS total       : {g.sum():+.1f} pts   (avg {g.mean():+.1f}/trade)")
    print(f"  NET total (cost {cost:.0f}): {n.sum():+.1f} pts   (avg {n.mean():+.1f}/trade)")
    print(f"  Net win rate      : {(n > 0).mean()*100:.0f}%  ({(n>0).sum()}/{len(n)})")
    print(f"  Best / worst (net): {n.max():+.1f} / {n.min():+.1f} pts")
    print(f"  Exit reasons      : {tr.exit_reason.value_counts().to_dict()}")
    print("  -- Long vs Short spread (net of cost) --")
    ls_rows = []
    for dirn in ["LONG_SPREAD", "SHORT_SPREAD"]:
        s = tr[tr.direction == dirn]
        if len(s):
            print(f"    {dirn}: {len(s)} trades | net {s.net_pnl_points.sum():+.1f} | "
                  f"win {(s.net_pnl_points > 0).mean()*100:.0f}% | avg hold {s.holding_days.mean():.1f}d")
            ls_rows.append({"direction": dirn, "trades": len(s),
                            "gross_pts": round(s.gross_pnl_points.sum(), 1),
                            "net_pts": round(s.net_pnl_points.sum(), 1),
                            "win_pct": round((s.net_pnl_points > 0).mean()*100, 0)})
    print("=" * 70)

    # profitable-only view
    winners = tr[tr.net_pnl_points > 0]
    winners.to_csv("backtest_daily_profitable_trades.csv", index=False)

    with pd.ExcelWriter("backtest_daily_report.xlsx", engine="openpyxl") as xl:
        summary = pd.DataFrame({
            "metric": ["Days", "Trades", "Avg hold (days)", "Gross total (pts)",
                       "Net total (pts)", "Net win rate %", "Best net", "Worst net"],
            "value": [len(df), len(tr), round(tr.holding_days.mean(), 1), round(g.sum(), 1),
                      round(n.sum(), 1), round((n > 0).mean()*100, 0), round(n.max(), 1), round(n.min(), 1)],
        })
        summary.to_excel(xl, sheet_name="Summary", index=False)
        pd.DataFrame(ls_rows).to_excel(xl, sheet_name="Long vs Short", index=False)
        tr.to_excel(xl, sheet_name="All Trades", index=False)
        winners.to_excel(xl, sheet_name="Profitable Trades", index=False)
        for ws in xl.book.worksheets:
            fmt(ws)
    print("Wrote backtest_daily_report.xlsx (+ backtest_daily_trades.csv, "
          "backtest_daily_profitable_trades.csv)")


if __name__ == "__main__":
    main()
