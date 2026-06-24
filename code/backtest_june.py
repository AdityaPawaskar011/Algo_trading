"""
backtest_june.py — MONTHLY (multi-day, position-held) backtest of the spread
strategy for a single month (default June 2026).

These are FUTURES with monthly expiry, so a position is carried OVERNIGHT and
held across days until the spread reverts (|z| <= EXIT), a stop/target hits, or
the monthly expiry approaches (MAX_HOLD) — NOT squared off intraday.

The daily series = history.csv (Yahoo daily closes) + the recent live-feed day
closes, so the z-score has the trailing lookback it needs. Only trades ENTERED
in the target month are reported.

Run from the feed_data/ folder:
    python ..\\code\\backtest_june.py
    python ..\\code\\backtest_june.py --month 2026-06 --cost-pts 35 --target 200 --maxhold 25
"""
import sys
import pandas as pd
from openpyxl.styles import Font, PatternFill

import backtest as bt

FEED_DAYS = [
    ("2026-06-18", "sensex_Nifty/sensex_data.csv",        "sensex_Nifty/nifty_data.csv"),
    ("2026-06-19", "sensex_data_2026-06-19_archived.csv", "nifty_data_2026-06-19_archived.csv"),
    ("2026-06-22", "sensex_data.csv",                     "nifty_data.csv"),
    ("2026-06-23", "sensex_today.csv",                    "nifty_today.csv"),
]


def arg(flag, default):
    return sys.argv[sys.argv.index(flag) + 1] if flag in sys.argv else default


def daily_close(path):
    df = pd.read_csv(path, usecols=["tick_time", "last_price"]).dropna()
    return float(df.iloc[-1]["last_price"])


def build_daily():
    h = pd.read_csv("history.csv")
    h["date"] = pd.to_datetime(h["date"])
    extra = []
    for d, sxf, nff in FEED_DAYS:
        try:
            extra.append({"date": pd.Timestamp(d),
                          "sensex_close": daily_close(sxf),
                          "nifty_close": daily_close(nff)})
        except Exception:
            pass
    if extra:
        h = pd.concat([h, pd.DataFrame(extra)], ignore_index=True)
    return h.drop_duplicates("date").sort_values("date").reset_index(drop=True)[
        ["date", "sensex_close", "nifty_close"]]


def fmt(ws):
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E78")
    ws.freeze_panes = "A2"


def main():
    month = arg("--month", "2026-06")
    cost  = float(arg("--cost-pts", "4"))
    # "long trade" = let the position run; do not cap winners at the tiny +30
    bt.PROFIT_TARGET = float(arg("--target", "200"))
    bt.MAX_HOLD      = int(arg("--maxhold", "25"))

    df = build_daily()
    print(f"Daily series: {len(df)} days ({df.date.min().date()} -> {df.date.max().date()})")
    print(f"Month: {month} | hold-to-reversion (multi-day) | entry +/-{bt.ENTRY} exit +/-{bt.EXIT} "
          f"SL -{bt.STOP_LOSS} TP +{bt.PROFIT_TARGET} max-hold {bt.MAX_HOLD}d | cost {cost:.0f} pts")

    tr = bt.backtest(df)
    if not len(tr):
        print("No trades generated at all.")
        return

    tr["entry_date"] = pd.to_datetime(tr["entry_date"])
    tr["exit_date"]  = pd.to_datetime(tr["exit_date"])
    jun = tr[tr["entry_date"].dt.strftime("%Y-%m") == month].copy()
    if not len(jun):
        print(f"No trades ENTERED in {month}.  "
              f"(Strategy entered trades on: {sorted(tr.entry_date.dt.strftime('%Y-%m').unique())[-6:]})")
        return

    jun = jun.rename(columns={"pnl_points": "gross_pnl_points"})
    jun["cost_points"]    = cost
    jun["net_pnl_points"] = (jun["gross_pnl_points"] - cost).round(2)
    jun["cum_net_points"] = jun["net_pnl_points"].cumsum().round(2)
    jun["result"] = jun["net_pnl_points"].apply(lambda p: "WIN" if p > 0 else "LOSS")
    cols = ["entry_date", "exit_date", "direction", "spread_type", "holding_days",
            "sensex_entry", "nifty_entry", "ratio", "spread_entry", "zscore_entry",
            "spread_exit", "zscore_exit", "exit_reason",
            "gross_pnl_points", "cost_points", "net_pnl_points", "cum_net_points", "result"]
    jun = jun[[c for c in cols if c in jun.columns]]
    jun["entry_date"] = jun["entry_date"].dt.strftime("%Y-%m-%d")
    jun["exit_date"]  = jun["exit_date"].dt.strftime("%Y-%m-%d")

    g, n = jun.gross_pnl_points, jun.net_pnl_points
    print("\n" + "=" * 68)
    print(f"  {month} MONTHLY (multi-day hold) SPREAD BACKTEST")
    print("=" * 68)
    print(jun[["entry_date", "exit_date", "direction", "holding_days",
               "zscore_entry", "zscore_exit", "exit_reason",
               "gross_pnl_points", "net_pnl_points"]].to_string(index=False))
    print("-" * 68)
    print(f"  Trades {len(jun)} | avg hold {jun.holding_days.mean():.1f} days "
          f"(max {jun.holding_days.max()})")
    print(f"  GROSS {g.sum():+.1f} | NET {n.sum():+.1f} pts | net win {(n > 0).mean()*100:.0f}%")
    ls_rows = []
    for dirn in ["LONG_SPREAD", "SHORT_SPREAD"]:
        s = jun[jun.direction == dirn]
        if len(s):
            print(f"  {dirn}: {len(s)} trades | net {s.net_pnl_points.sum():+.1f} | "
                  f"win {(s.net_pnl_points > 0).mean()*100:.0f}%")
            ls_rows.append({"direction": dirn, "trades": len(s),
                            "gross_pts": round(s.gross_pnl_points.sum(), 1),
                            "net_pts": round(s.net_pnl_points.sum(), 1),
                            "win_pct": round((s.net_pnl_points > 0).mean()*100, 0)})
    print("=" * 68)

    out = f"backtest_{month.replace('-', '_')}_monthly.xlsx"
    summary = pd.DataFrame({
        "metric": ["Month", "Trade type", "Trades", "Avg hold (days)", "Max hold (days)",
                   "Gross total (pts)", "Cost/trade (pts)", "Net total (pts)",
                   "Net win rate %", "Best net", "Worst net"],
        "value": [month, "multi-day (held to reversion / expiry)", len(jun),
                  round(jun.holding_days.mean(), 1), int(jun.holding_days.max()),
                  round(g.sum(), 1), cost, round(n.sum(), 1),
                  round((n > 0).mean()*100, 0), round(n.max(), 1), round(n.min(), 1)],
    })
    with pd.ExcelWriter(out, engine="openpyxl") as xl:
        summary.to_excel(xl, sheet_name="Summary", index=False)
        if ls_rows:
            pd.DataFrame(ls_rows).to_excel(xl, sheet_name="Long vs Short", index=False)
        jun.to_excel(xl, sheet_name=f"{month} Trades", index=False)
        for ws in xl.book.worksheets:
            fmt(ws)
    print(f"Wrote {out}  ({len(jun)} {month} trades)")


if __name__ == "__main__":
    main()
