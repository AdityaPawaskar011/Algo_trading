"""
cost_backtester.py — realistic, COST-AWARE backtester for the SENSEX-NIFTY spread.

Net-of-cost is the headline. Gross is shown only for contrast. A single day is
not proof of an edge, and a config only "counts" if it is net-positive AND took
>= 5 trades — otherwise any profit is a 1-2 trade fluke.

Run from feed_data/:
    python ..\\code\\cost_backtester.py --sensex sensex_today.csv --nifty nifty_today.csv --tag 2026_06_24

Outputs one .xlsx (Summary, Trades 1min, Trades per-second, All Configs,
Profitable Configs, Price Data, Multi-Lot) and prints the Summary.
"""
import sys
import itertools
import numpy as np
import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import LineChart, Reference

# ══════════════════════════ CONFIG ══════════════════════════
CFG = {
    # ── data / session ──
    "session": "09:15-15:30 IST",
    # ── strategy defaults (the headline 1-min run uses these) ──
    "ratio_win": 60, "spread_win": 15,
    "entry_z": 2.0, "exit_z": 0.7,
    "profit_target": 30, "stop_loss": 100, "max_hold": 15,   # max_hold in BARS
    "bars": ["1min", "3min"],                                 # resample sizes for the sweep
    # ── transaction-cost model (per 1-lot pair, per round-trip) ──
    "per_leg_notional": 1_500_000,   # Rs per leg
    "stt_rate":        0.0,       # 0.02% on the SELL side (2 sells / round-trip)
    "brokerage_order": 20,           # Rs per executed order (4 orders / round-trip)
    "exchange_rate":   0.0,      # ~0.002% of turnover (blended NSE/BSE F&O)
    "sebi_rate":       0.0,     # ~0.0001% of turnover
    "gst_rate":        0.18,         # on brokerage+exchange+sebi
    "stamp_rate":      0.0,      # ~0.002% on the BUY side (2 buys / round-trip)
    "slippage_points": 0,            # extra points lost crossing the book, round-trip
    "point_value_rs":  20,           # Rs per spread point per 1-lot pair (SENSEX fut lot 20)
    "margin_per_lot":  150_000,      # approx span+exposure margin per 1-lot pair
}
# sweep grid
GRID = {"entry_z": [2.0, 2.5, 3.0], "exit_z": [0.5, 0.7, 1.0],
        "profit_target": [30, 60, 120], "stop_loss": [100, 150],
        "max_hold": [15, 30], "bar": ["1min", "3min"]}
LOTS = [1, 2, 5, 10, 20]
GREEN = PatternFill("solid", fgColor="C6EFCE"); RED = PatternFill("solid", fgColor="FFC7CE")
HEAD  = PatternFill("solid", fgColor="1F4E78"); BANNER = PatternFill("solid", fgColor="FFF2CC")


def arg(flag, d): return sys.argv[sys.argv.index(flag) + 1] if flag in sys.argv else d


# ── cost model ──
def round_trip_cost():
    c = CFG
    notional, turnover = c["per_leg_notional"], c["per_leg_notional"] * 4
    stt   = c["stt_rate"] * notional * 2
    brok  = c["brokerage_order"] * 4
    exch  = c["exchange_rate"] * turnover
    sebi  = c["sebi_rate"] * turnover
    gst   = c["gst_rate"] * (brok + exch + sebi)
    stamp = c["stamp_rate"] * notional * 2
    rupee = stt + brok + exch + sebi + gst + stamp
    pts   = rupee / c["point_value_rs"] + c["slippage_points"]
    return pts, {"STT": stt, "brokerage": brok, "exchange": exch, "sebi": sebi,
                 "gst": gst, "stamp": stamp, "rupee_total": round(rupee, 1),
                 "slippage_pts": c["slippage_points"], "cost_points": round(pts, 1)}


# ── data ──
def load(sensex_csv, nifty_csv):
    sx = (pd.read_csv(sensex_csv, usecols=["tick_time", "last_price"]).dropna()
            .drop_duplicates("tick_time").rename(columns={"last_price": "sensex"}))
    nf = (pd.read_csv(nifty_csv, usecols=["tick_time", "last_price"]).dropna()
            .drop_duplicates("tick_time").rename(columns={"last_price": "nifty"}))
    df = pd.merge(sx, nf, on="tick_time", how="inner")
    df["time"] = pd.to_datetime(df["tick_time"])
    return df.sort_values("time")[["time", "sensex", "nifty"]].reset_index(drop=True)


def resample(df, rule):
    if rule in ("1s", "tick"):
        return df.copy()
    out = df.set_index("time").resample(rule).last().dropna().reset_index()
    return out


# ── strategy ──
def signals(bars, ratio_win, spread_win):
    d = bars.copy()
    d["ratio"]  = (d.sensex / d.nifty).rolling(ratio_win).mean()
    d["spread"] = d.sensex - d.ratio * d.nifty
    d["sma"]    = d.spread.rolling(spread_win).mean()
    d["ssd"]    = d.spread.rolling(spread_win).std()
    d["z"]      = (d.spread - d.sma) / d.ssd
    d["day"]    = d.time.dt.date
    return d


def backtest(bars, p, cost_pts):
    """One position at a time. Exit priority: stop-loss, profit-target,
    z-reversion, max-hold, end-of-day."""
    s = signals(bars, p["ratio_win"], p["spread_win"]).reset_index(drop=True)
    n = len(s); start = max(p["ratio_win"], p["spread_win"]); pos = None; out = []
    for i in range(start, n):
        r = s.iloc[i]; z = r.z
        if np.isnan(z):
            continue
        eod = (i == n - 1) or (s.iloc[i + 1].day != r.day)
        if pos is not None:
            held = i - pos["i"]
            live = (r.spread - pos["spread"]) if pos["dir"] == "LONG" else (pos["spread"] - r.spread)
            why = ""
            if   live <= -p["stop_loss"]:        why = "stop_loss"
            elif live >=  p["profit_target"]:    why = "profit_target"
            elif pos["dir"] == "LONG"  and z >= -p["exit_z"]: why = "reverted"
            elif pos["dir"] == "SHORT" and z <=  p["exit_z"]: why = "reverted"
            elif held >= p["max_hold"]:          why = "max_hold"
            elif eod:                            why = "eod"
            if why:
                out.append({"direction": pos["dir"], "entry_time": pos["t"], "exit_time": r.time,
                            "entry_sensex": pos["sx"], "entry_nifty": pos["nf"],
                            "spread_entry": round(pos["spread"], 2), "zscore_entry": round(pos["z"], 3),
                            "spread_exit": round(r.spread, 2), "zscore_exit": round(z, 3),
                            "holding_bars": held, "exit_reason": why,
                            "gross_pnl_points": round(live, 2)})
                pos = None
        if pos is None and not eod:
            if z <= -p["entry_z"]:
                pos = {"dir": "LONG", "i": i, "t": r.time, "sx": r.sensex, "nf": r.nifty, "spread": r.spread, "z": z}
            elif z >= p["entry_z"]:
                pos = {"dir": "SHORT", "i": i, "t": r.time, "sx": r.sensex, "nf": r.nifty, "spread": r.spread, "z": z}
    t = pd.DataFrame(out)
    if len(t):
        t["cost_points"]    = cost_pts
        t["net_pnl_points"] = (t.gross_pnl_points - cost_pts).round(2)
        t["cum_net_points"] = t.net_pnl_points.cumsum().round(2)
        t["result"]         = np.where(t.net_pnl_points > 0, "WIN", "LOSS")
    return t


# ── metrics (gross AND net) ──
def metrics(t):
    if not len(t):
        return {}
    g, nt = t.gross_pnl_points, t.net_pnl_points
    gp, gl = g[g > 0].sum(), g[g < 0].sum()
    nw, nl = nt[nt > 0], nt[nt < 0]
    dd = (t.cum_net_points - t.cum_net_points.cummax()).min()
    # consecutive
    mcw = mcl = cw = cl = 0
    for v in nt:
        if v > 0: cw += 1; cl = 0
        else:     cl += 1; cw = 0
        mcw, mcl = max(mcw, cw), max(mcl, cl)
    hold_min = (pd.to_datetime(t.exit_time) - pd.to_datetime(t.entry_time)).dt.total_seconds() / 60
    return {
        "trades": len(t), "wins": int((nt > 0).sum()), "losses": int((nt < 0).sum()),
        "win_rate_pct": round((nt > 0).mean() * 100, 1),
        "gross_total_pts": round(g.sum(), 1), "net_total_pts": round(nt.sum(), 1),
        "gross_profit_pts": round(gp, 1), "gross_loss_pts": round(gl, 1),
        "profit_factor_gross": round(gp / abs(gl), 2) if gl else float("inf"),
        "expectancy_net_pts": round(nt.mean(), 2),
        "avg_win_net": round(nw.mean(), 2) if len(nw) else 0,
        "avg_loss_net": round(nl.mean(), 2) if len(nl) else 0,
        "payoff_ratio": round(nw.mean() / abs(nl.mean()), 2) if len(nw) and len(nl) and nl.mean() else 0,
        "largest_win_net": round(nt.max(), 1), "largest_loss_net": round(nt.min(), 1),
        "max_drawdown_net_pts": round(dd, 1),
        "sharpe_per_trade": round(nt.mean() / nt.std(), 2) if nt.std() else 0,
        "max_consec_wins": mcw, "max_consec_losses": mcl,
        "median_hold_min": round(hold_min.median(), 1), "max_hold_min": round(hold_min.max(), 1),
        "long_net_pts": round(nt[t.direction == "LONG"].sum(), 1),
        "short_net_pts": round(nt[t.direction == "SHORT"].sum(), 1),
    }


# ── sweep ──
def sweep(by_bar, cost_pts):
    rows = []
    base = dict(ratio_win=CFG["ratio_win"], spread_win=CFG["spread_win"])
    for entry, exit_, tp, sl, mh, bar in itertools.product(
            GRID["entry_z"], GRID["exit_z"], GRID["profit_target"],
            GRID["stop_loss"], GRID["max_hold"], GRID["bar"]):
        bars = by_bar[bar]
        if len(bars) <= CFG["ratio_win"] + 2:
            continue
        p = {**base, "entry_z": entry, "exit_z": exit_, "profit_target": tp,
             "stop_loss": sl, "max_hold": mh}
        t = backtest(bars, p, cost_pts)
        if not len(t):
            continue
        g, nt = t.gross_pnl_points, t.net_pnl_points
        rows.append({"bar": bar, "entry_z": entry, "exit_z": exit_, "profit_target": tp,
                     "stop_loss": sl, "max_hold": mh, "trades": len(t),
                     "gross_pts": round(g.sum(), 1), "net_pts": round(nt.sum(), 1),
                     "net_win_pct": round((nt > 0).mean() * 100, 1),
                     "avg_net_pts": round(nt.mean(), 2),
                     "best_pts": round(g.max(), 1), "worst_pts": round(g.min(), 1)})
    return pd.DataFrame(rows).sort_values("net_pts", ascending=False).reset_index(drop=True)


def best_cfg(df_sweep, min_trades=5):
    pool = df_sweep[df_sweep.trades >= min_trades]
    return (pool.iloc[0] if len(pool) else None)


def main():
    sx_csv = arg("--sensex", "sensex_today.csv")
    nf_csv = arg("--nifty",  "nifty_today.csv")
    tag    = arg("--tag", "report")
    out    = arg("--out", f"cost_backtest_{tag}.xlsx")

    cost_pts, cost_breakdown = round_trip_cost()
    df = load(sx_csv, nf_csv)
    by_bar = {"1s": df, "1min": resample(df, "1min"), "3min": resample(df, "3min")}
    day0 = df.time.iloc[0].date()
    print(f"Data: {len(df)} ticks  {df.time.min()} -> {df.time.max()}")
    print(f"Round-trip cost: {cost_pts:.1f} pts (Rs {cost_breakdown['rupee_total']}/lot)  {cost_breakdown}")

    base = dict(ratio_win=CFG["ratio_win"], spread_win=CFG["spread_win"], entry_z=CFG["entry_z"],
                exit_z=CFG["exit_z"], profit_target=CFG["profit_target"],
                stop_loss=CFG["stop_loss"], max_hold=CFG["max_hold"])
    t1   = backtest(by_bar["1min"], base, cost_pts)   # headline
    tsec = backtest(by_bar["1s"],   base, cost_pts)   # contrast (noise)
    m1   = metrics(t1)

    sw = sweep(by_bar, cost_pts)
    profitable = sw[sw.net_pts > 0]
    robust = profitable[profitable.trades >= 5]
    n_cfg, n_prof, n_robust = len(sw), len(profitable), len(robust)

    # walk-forward: morning train / afternoon test on 1-min
    bars1 = by_bar["1min"]; mid = bars1.time.iloc[len(bars1)//2]
    am, pm = bars1[bars1.time <= mid], bars1[bars1.time > mid]
    wf_train_net = wf_test_net = None; wf_cfg = None
    if len(am) > CFG["ratio_win"] + 5 and len(pm) > CFG["ratio_win"] + 5:
        sw_am = sweep({"1min": am, "3min": am}, cost_pts)
        bc = best_cfg(sw_am, min_trades=3)
        if bc is not None:
            wf_cfg = bc
            wf_train_net = bc.net_pts
            p = {**base, "entry_z": bc.entry_z, "exit_z": bc.exit_z, "profit_target": bc.profit_target,
                 "stop_loss": bc.stop_loss, "max_hold": bc.max_hold}
            tt = backtest(pm, p, cost_pts)
            wf_test_net = round(tt.net_pnl_points.sum(), 1) if len(tt) else 0.0

    # break-even cost (headline 1-min, default params): gross avg/trade
    be = round(t1.gross_pnl_points.mean(), 1) if len(t1) else 0.0

    # multi-lot scaling (headline net)
    pv = CFG["point_value_rs"]
    lot_rows = [{"lots": L, "gross_Rs": round(m1.get("gross_total_pts", 0) * pv * L),
                 "net_Rs": round(m1.get("net_total_pts", 0) * pv * L),
                 "margin_Rs": CFG["margin_per_lot"] * L} for L in LOTS]

    net1 = m1.get("net_total_pts", 0)
    verdict = "PROFITABLE after costs" if (net1 > 0 and n_robust > 0) else "NOT profitable after costs"

    # ── Summary (metric / value / what_it_means) ──
    S = [
        ("Date", str(day0), "Single trading day of per-second SENSEX/NIFTY data"),
        ("Session", CFG["session"], ""),
        ("Round-trip cost (points)", round(cost_pts, 1),
         f"STT+brokerage+exch+GST+stamp (Rs{cost_breakdown['rupee_total']}/lot) + {CFG['slippage_points']} slippage"),
        ("Bar size (headline)", "1-min", "Resampled; raw 1-second ticks are noise (shown only for contrast)"),
        ("-- HEADLINE (1-min, default params) --", "", ""),
        ("Trades", m1.get("trades", 0), "One position at a time"),
        ("Win rate %", m1.get("win_rate_pct", 0), "Share of net-positive trades"),
        ("GROSS P&L (pts)", m1.get("gross_total_pts", 0), "Before costs - NOT the conclusion"),
        ("NET P&L (pts)", net1, "AFTER costs - THE conclusion"),
        ("NET P&L (Rs/lot)", round(net1 * pv), f"at Rs{pv}/point/lot"),
        ("Profit factor (gross)", m1.get("profit_factor_gross", 0), "Gross profit / gross loss"),
        ("Expectancy net (pts/trade)", m1.get("expectancy_net_pts", 0), "Avg net P&L per trade"),
        ("Payoff ratio (net)", m1.get("payoff_ratio", 0), "Avg win / avg loss"),
        ("Max drawdown net (pts)", m1.get("max_drawdown_net_pts", 0), "Worst peak-to-trough on net equity"),
        ("Sharpe (per-trade)", m1.get("sharpe_per_trade", 0), "Mean/Std of per-trade net"),
        ("Max consec wins / losses", f"{m1.get('max_consec_wins',0)} / {m1.get('max_consec_losses',0)}", ""),
        ("Median / max hold (min)", f"{m1.get('median_hold_min',0)} / {m1.get('max_hold_min',0)}", ""),
        ("LONG net / SHORT net (pts)", f"{m1.get('long_net_pts',0)} / {m1.get('short_net_pts',0)}", "P&L split by direction"),
        ("-- ROBUSTNESS --", "", ""),
        ("Configs swept", n_cfg, "entry/exit/target/stop/hold x bar (1m,3m)"),
        ("Net-positive configs", n_prof, "How many made money net of cost"),
        ("ROBUST configs (net>0 & >=5 trades)", n_robust,
         "If 0, any profit is a 1-2 trade fluke / curve-fit"),
        ("Walk-forward train net (pts)", wf_train_net if wf_train_net is not None else "n/a",
         "Best config fit on the MORNING"),
        ("Walk-forward test net (pts)", wf_test_net if wf_test_net is not None else "n/a",
         "Same config on the UNSEEN AFTERNOON - this is what matters"),
        ("Break-even cost (pts/trade)", be, "Cost at which net hits zero; compare to the cost above"),
        ("-- VERDICT --", "", ""),
        ("VERDICT", verdict,
         f"Edge ~{be:.0f} pts/trade vs cost ~{cost_pts:.0f} pts/trade" if "NOT" in verdict
         else "Net positive with a robust config - still needs multi-day validation"),
        ("Caveat", "One day is not proof", "Validate across many days + walk-forward before trading real money"),
    ]
    summary = pd.DataFrame(S, columns=["metric", "value", "what_it_means"])

    # ── write workbook ──
    price = by_bar["1min"].rename(columns={"time": "time", "sensex": "sensex", "nifty": "nifty"})
    with pd.ExcelWriter(out, engine="openpyxl") as xl:
        summary.to_excel(xl, sheet_name="Summary", index=False)
        (t1 if len(t1) else pd.DataFrame()).to_excel(xl, sheet_name="Trades (1min)", index=False)
        (tsec if len(tsec) else pd.DataFrame()).to_excel(xl, sheet_name="Trades (per-second)", index=False)
        sw.to_excel(xl, sheet_name="All Configs", index=False)
        profitable.to_excel(xl, sheet_name="Profitable Configs", index=False)
        pd.DataFrame(lot_rows).to_excel(xl, sheet_name="Multi-Lot", index=False)
        price.to_excel(xl, sheet_name="Price Data", index=False)
        wb = xl.book
        # formatting
        for name in wb.sheetnames:
            ws = wb[name]
            for c in ws[1]:
                c.font = Font(bold=True, color="FFFFFF"); c.fill = HEAD
            ws.freeze_panes = "A2"
        # colour P&L cells
        for name, mc in [("Trades (1min)", ["gross_pnl_points", "net_pnl_points", "cum_net_points"]),
                         ("Trades (per-second)", ["gross_pnl_points", "net_pnl_points", "cum_net_points"]),
                         ("All Configs", ["net_pts"]), ("Profitable Configs", ["net_pts"]),
                         ("Multi-Lot", ["net_Rs"])]:
            ws = wb[name]; hdr = {c.value: c.column for c in ws[1]}
            for col in mc:
                ci = hdr.get(col)
                if not ci:
                    continue
                for r in range(2, ws.max_row + 1):
                    cell = ws.cell(row=r, column=ci)
                    try:
                        v = float(cell.value)
                    except (TypeError, ValueError):
                        continue
                    if v > 0: cell.fill = GREEN
                    elif v < 0: cell.fill = RED
        # equity curve chart on Trades (1min)
        ws = wb["Trades (1min)"]
        if len(t1) and "cum_net_points" in [c.value for c in ws[1]]:
            ci = [c.value for c in ws[1]].index("cum_net_points") + 1
            ch = LineChart(); ch.title = "Net equity curve (1-min)"; ch.y_axis.title = "cum net pts"
            data = Reference(ws, min_col=ci, min_row=1, max_row=ws.max_row)
            ch.add_data(data, titles_from_data=True)
            ws.add_chart(ch, "T2")

    print("\n" + "=" * 70)
    print(f"  VERDICT: {verdict}")
    print("=" * 70)
    print(summary.to_string(index=False))
    print("=" * 70)
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
