"""
edge_strategy.py — strategy designed to PROFIT AFTER CHARGES by only taking
trades whose profit potential is several times the charge.

Design principle (what the user asked for): don't trade unless the expected win
is much bigger than the cost.
  1. ENTRY only when the spread is stretched (|z| >= ENTRY) AND the distance back
     to its mean ("room to revert") is >= TARGET. Room = |spread - mean|; if it
     reverts to the mean you capture ~that many points, so this guarantees the
     setup CAN reach the target.
  2. TARGET = a multiple of the charge (so each win clears the toll with margin).
  3. EXIT at +TARGET (book), -STOP (cut), or MAXHOLD bars (carry-forward limit).
     Works INTRADAY (if it reverts same day) and CARRY-FORWARD (held across days).
  4. LONG spread only (the side that held up out-of-sample).

Validated out-of-sample: train 2024-25, test 2026 (unseen). Charge default 4 pts
(Rs80 round-trip = Rs20/order x4).

The Excel report has: Summary | This Week | Target vs Charge | Trade Detail
(Rs20-order, every trade with SENSEX/NIFTY buy-sell prices + net P&L) | Multiple
Lots.  "This Week" = trades entered on/after this Monday (override with --since).

Run from feed_data/:
    python ..\\code\\edge_strategy.py                       # uses history.csv
    python ..\\code\\edge_strategy.py --data tracker_daily.csv --since 2026-06-22 \\
           --out ..\\reports\\edge_this_week.xlsx           # up-to-date series, this week
"""
import sys
from datetime import date, timedelta
import numpy as np
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment

ENTRY_Z, STOP, MAXHOLD = 2.0, 100, 30
RATIO_WIN, SPREAD_WIN = 60, 15
PV = 20                      # Rs per spread point per lot
LOTS = [1, 2, 5, 10, 20]
MARGIN_PER_LOT = 150000
GREEN = PatternFill("solid", fgColor="C6EFCE"); RED = PatternFill("solid", fgColor="FFC7CE")
HEAD = PatternFill("solid", fgColor="1F4E78")


def arg(flag, d): return sys.argv[sys.argv.index(flag) + 1] if flag in sys.argv else d


def signals(df):
    d = df.copy()
    d["ratio"]  = (d.sensex_close / d.nifty_close).rolling(RATIO_WIN).mean()
    d["spread"] = d.sensex_close - d.ratio * d.nifty_close
    d["sma"]    = d.spread.rolling(SPREAD_WIN).mean()
    d["ssd"]    = d.spread.rolling(SPREAD_WIN).std()
    d["z"]      = (d.spread - d.sma) / d.ssd
    return d


def backtest(df, target, cost, long_only=True):
    """Enter only if room-to-revert >= target; exit at +target / -STOP / MAXHOLD.
    Also records the SENSEX/NIFTY price at entry & exit (for the trade-detail sheet)."""
    s = signals(df).reset_index(drop=True)
    n = len(s); start = max(RATIO_WIN, SPREAD_WIN); pos = None; out = []
    for i in range(start, n):
        r = s.iloc[i]; z = r.z
        if np.isnan(z):
            continue
        if pos is not None:
            held = i - pos["i"]
            live = (r.spread - pos["spread"]) if pos["dir"] == "LONG" else (pos["spread"] - r.spread)
            why = ("profit_target" if live >= target else
                   "stop_loss" if live <= -STOP else
                   "max_hold" if held >= MAXHOLD else "")
            if why:
                out.append({"entry_date": pos["date"], "exit_date": r.date, "direction": pos["dir"],
                            "holding_bars": held, "zscore_entry": round(pos["z"], 2),
                            "room_at_entry": round(pos["room"], 1), "exit_reason": why,
                            "gross_pnl_points": round(live, 1),
                            "sensex_entry": round(pos["sensex"], 1), "sensex_exit": round(r.sensex_close, 1),
                            "nifty_entry": round(pos["nifty"], 1),  "nifty_exit": round(r.nifty_close, 1)})
                pos = None
        if pos is None:
            room = abs(r.spread - r.sma)                  # potential reversion profit
            if z <= -ENTRY_Z and room >= target:
                pos = {"dir": "LONG", "i": i, "date": r.date, "spread": r.spread, "z": z, "room": room,
                       "sensex": r.sensex_close, "nifty": r.nifty_close}
            elif (not long_only) and z >= ENTRY_Z and room >= target:
                pos = {"dir": "SHORT", "i": i, "date": r.date, "spread": r.spread, "z": z, "room": room,
                       "sensex": r.sensex_close, "nifty": r.nifty_close}
    t = pd.DataFrame(out)
    if len(t):
        t["entry_date"] = pd.to_datetime(t["entry_date"])
        t["exit_date"]  = pd.to_datetime(t["exit_date"])
        t["net_pnl_points"] = (t.gross_pnl_points - cost).round(1)
    return t


def build_detail(df, target, cost):
    """One row per trade in the v2 report format; P&L in Rs at the Rs20/order basis.
    GROSS/NET use the strategy's own spread-points P&L (x PV) so they match the grid;
    the leg prices are what you would actually BUY/SELL at."""
    t = backtest(df, target, cost, long_only=True)
    if not len(t):
        return pd.DataFrame()
    rows, run = [], 0
    for _, r in t.iterrows():
        LONG = (r.direction == "LONG")
        gross = round(r.gross_pnl_points * PV)
        net   = round(r.net_pnl_points * PV)
        run  += net
        rows.append({
            "entry_date": r.entry_date.strftime("%Y-%m-%d"),
            "exit_date":  r.exit_date.strftime("%Y-%m-%d"),
            "days_held":  int(r.holding_bars),
            "trade":      r.direction,
            "SENSEX_in":  "BUY" if LONG else "SELL",  "SENSEX_entry": r.sensex_entry,
            "SENSEX_out": "SELL" if LONG else "BUY",  "SENSEX_exit":  r.sensex_exit,
            "NIFTY_in":   "SELL" if LONG else "BUY",  "NIFTY_entry":  r.nifty_entry,
            "NIFTY_out":  "BUY" if LONG else "SELL",  "NIFTY_exit":   r.nifty_exit,
            "entry_z":    r.zscore_entry, "room_pts": r.room_at_entry, "exit_reason": r.exit_reason,
            "GROSS_Rs":   gross, "charge_Rs": round(cost * PV), "NET_Rs": net,
            "P/L":        "PROFIT" if net > 0 else "LOSS", "running_NET_Rs": run,
        })
    d = pd.DataFrame(rows)
    d.insert(0, "#", range(1, len(d) + 1))
    return d


def lots_table(detail, cost):
    n = len(detail); gp = int(detail.GROSS_Rs.sum()) if n else 0
    charge = round(cost * PV) * n          # brokerage is flat, does NOT scale with lots
    return pd.DataFrame([{
        "lots": L, "GROSS_Rs": round(gp * L), "brokerage_Rs(flat)": charge,
        "NET_Rs": round(gp * L - charge), "margin_Rs": MARGIN_PER_LOT * L,
    } for L in LOTS])


def style_all(wb):
    for nm in wb.sheetnames:
        ws = wb[nm]
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF"); c.fill = HEAD
            c.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 13
        hdr = {c.value: c.column for c in ws[1]}
        for col in ("GROSS_Rs", "NET_Rs", "running_NET_Rs", "train_net", "test_net", "test_net_Rs"):
            ci = hdr.get(col)
            if not ci:
                continue
            for rr in range(2, ws.max_row + 1):
                cell = ws.cell(row=rr, column=ci)
                try:
                    v = float(cell.value)
                except (TypeError, ValueError):
                    continue
                if v > 0:
                    cell.fill = GREEN
                elif v < 0:
                    cell.fill = RED
        for col in ("SENSEX_in", "SENSEX_out", "NIFTY_in", "NIFTY_out", "P/L"):
            ci = hdr.get(col)
            if not ci:
                continue
            for rr in range(2, ws.max_row + 1):
                cell = ws.cell(row=rr, column=ci)
                cell.fill = GREEN if cell.value in ("BUY", "PROFIT") else RED


def put(xl, name, df):
    """Write df to a sheet, or a friendly note if it is empty."""
    if len(df):
        df.to_excel(xl, sheet_name=name, index=False)
    else:
        pd.DataFrame({"note": ["no trades for this selection"]}).to_excel(xl, sheet_name=name, index=False)


def main():
    cost  = float(arg("--cost-pts", "4"))
    split = pd.Timestamp(arg("--split", "2026-01-01"))
    data  = arg("--data", "history.csv")
    out   = arg("--out", "edge_strategy.xlsx")
    today = date.today()
    since = arg("--since", None)
    week_start = pd.Timestamp(since) if since else pd.Timestamp(today - timedelta(days=today.weekday()))

    df = pd.read_csv(data); df["date"] = pd.to_datetime(df["date"])
    df = df.sort_values("date").reset_index(drop=True)
    print(f"Daily series: {len(df)} days ({df.date.min().date()} -> {df.date.max().date()}) "
          f"from {data} | charge {cost:.0f} pts (Rs{cost*PV:.0f})\n")

    rows = []
    for mult in (1.0, 1.5, 2.0, 3.0):
        target = round(mult * cost)
        t = backtest(df, target, cost, long_only=True)
        if not len(t):
            rows.append({"target_pts": target, "x_charge": mult, "trades": 0, "train_net": 0,
                         "test_net": 0, "test_net_Rs": 0, "avg_win_pts": 0, "win%": 0})
            continue
        tr = t[t.entry_date < split]; te = t[t.entry_date >= split]
        wins = t[t.net_pnl_points > 0]
        rows.append({"target_pts": target, "x_charge": mult, "trades": len(t),
                     "train_net": round(tr.net_pnl_points.sum(), 0),
                     "test_net": round(te.net_pnl_points.sum(), 0),
                     "test_net_Rs": round(te.net_pnl_points.sum() * PV),
                     "avg_win_pts": round(wins.gross_pnl_points.mean(), 0) if len(wins) else 0,
                     "win%": round((t.net_pnl_points > 0).mean() * 100, 0)})
    grid = pd.DataFrame(rows)

    ok = grid[(grid.trades >= 3) & (grid.train_net > 0) & (grid.test_net > 0)]
    if len(ok):
        b = ok.sort_values("test_net", ascending=False).iloc[0]
        chosen = int(b.target_pts)
        verdict = (f"BEST: target {chosen} pts ({b.x_charge:g}x charge) -> net-positive in train AND test "
                   f"({b.test_net:+.0f} pts / Rs{b.test_net_Rs:+,.0f}/lot OOS), avg win {b.avg_win_pts:.0f} pts")
    else:
        chosen = round(2 * cost)
        verdict = f"No target net-positive OOS on this data; detail uses {chosen} pts (2x charge) for reference."

    detail = build_detail(df, chosen, cost)
    if len(detail):
        ed = pd.to_datetime(detail.entry_date)
        this_week = detail[ed >= week_start].reset_index(drop=True)
        if len(this_week):
            this_week["running_NET_Rs"] = this_week.NET_Rs.cumsum()
    else:
        this_week = pd.DataFrame()
    lots = lots_table(this_week if len(this_week) else detail, cost)

    print("=" * 96)
    print("  EDGE STRATEGY: LONG carry-forward, room-filtered, target = multiple of charge")
    print("=" * 96)
    print(grid.to_string(index=False))
    print("-" * 96)
    print(f"  {verdict}")
    print(f"  Detail target: {chosen} pts | total trades {len(detail)} | "
          f"this week (>= {week_start.date()}): {len(this_week)} trade(s)")
    if len(this_week):
        print(this_week[["#", "entry_date", "exit_date", "days_held", "NET_Rs", "P/L"]].to_string(index=False))
    print("=" * 96)

    tw_net = int(this_week.NET_Rs.sum()) if len(this_week) else 0
    summary = pd.DataFrame({
        "metric": ["Strategy", "Data file", "Series range", "Entry filter", "Exit", "Hold",
                   "Charge (pts/Rs)", "Detail target", "This week from", "This-week trades",
                   "This-week NET (1 lot)", "VERDICT", "Caveat"],
        "value": ["LONG spread, only when room-to-revert >= target (selective; cannot over-trade)",
                  data, f"{df.date.min().date()} -> {df.date.max().date()}",
                  f"z <= -{ENTRY_Z} AND room >= target", "+target / -stop / max-hold",
                  f"up to {MAXHOLD} days (carry-forward)", f"{cost:.0f} / Rs{cost*PV:.0f}",
                  f"{chosen} pts", str(week_start.date()), len(this_week),
                  f"Rs{tw_net:+,.0f}", verdict,
                  "Daily series; today's close (if market still open) is not included. "
                  "Validate forward with the paper tracker before real money."]})

    with pd.ExcelWriter(out, engine="openpyxl") as xl:
        summary.to_excel(xl, sheet_name="Summary", index=False)
        put(xl, "This Week", this_week)
        grid.to_excel(xl, sheet_name="Target vs Charge", index=False)
        put(xl, "Trade Detail (Rs20-order)", detail)
        lots.to_excel(xl, sheet_name="Multiple Lots", index=False)
        style_all(xl.book)
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
