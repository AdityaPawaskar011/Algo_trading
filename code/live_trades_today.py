r"""
live_trades_today.py — build an Excel of today's intraday spread trades computed
DIRECTLY FROM THE LIVE FEED (sensex_today.csv / nifty_today.csv), not from the
executed paper_trades.csv. It runs the exact intraday strategy on the per-second
feed, so the report is derived from the raw ticks and works with or without the
paper trader running.

Strategy (same engine as the v2 / oldfolder reports — backtest.compute_signals +
backtest, on PER-SECOND ticks, no resampling):
    ratio = rolling Sensex/Nifty (60 ticks); spread = Sensex - ratio*Nifty;
    z over 15 ticks; enter |z|>=2, exit reversion |z|<=0.7 / stop -100 /
    target +30 / max-hold 15.
This reproduces the live --intraday trader closely (the live trader uses a
400-tick rolling buffer in compute_intraday, so counts differ by a few trades).

Cost basis: Rs20/order x4 = Rs80/round-trip. P&L = Rs20 per spread point per lot.

Output: reports/live_trades_<date>.xlsx  (Summary | Live Trades | Multiple Lots)

Modes:
    one-shot (default) — build once from the feed as it currently stands:
        ..\log_tradingVenv\Scripts\python.exe ..\code\live_trades_today.py
    LIVE watch         — rebuild on EVERY new trade (polls every N sec), until
                         15:30 IST or Ctrl+C. Keep the Excel CLOSED so it can
                         overwrite; it skips+retries while the file is open:
        ..\log_tradingVenv\Scripts\python.exe ..\code\live_trades_today.py --watch
        ..\log_tradingVenv\Scripts\python.exe ..\code\live_trades_today.py --watch 10

Run from feed_data/ (so it finds the feed CSVs).
"""
import os
import sys
import time
from datetime import date, datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment

import backtest as bt
from backtest_intraday import load_series

# ── strategy params: EXACTLY the v2 intraday settings ──
bt.ENTRY, bt.EXIT, bt.STOP_LOSS, bt.PROFIT_TARGET, bt.MAX_HOLD, bt.RATIO_LOOKBACK, bt.LOOKBACK = \
    2.0, 0.7, 100, 30, 15, 60, 15

PV        = 20            # Rs per spread point per lot
CHARGE    = 20 * 4        # Rs80 round-trip (Rs20/order x 4 orders)
LOTS      = [1, 2, 5, 10, 20]
MARGIN_PER_LOT = 150000
EOD_HH, EOD_MM = 15, 30   # stop the watcher at market close (IST)

GREEN = PatternFill("solid", fgColor="C6EFCE")
RED   = PatternFill("solid", fgColor="FFC7CE")
HEAD  = PatternFill("solid", fgColor="1F4E78")

HERE      = os.path.dirname(os.path.abspath(__file__))
FEED_DIR  = os.path.join(HERE, "..", "feed_data")
OUT_DIR   = os.path.join(HERE, "..", "reports")


def arg(flag, default):
    return sys.argv[sys.argv.index(flag) + 1] if flag in sys.argv else default


def resolve(path):
    """Use path as-is if it exists (run from feed_data/), else look in feed_data/."""
    if os.path.exists(path):
        return path
    alt = os.path.join(FEED_DIR, os.path.basename(path))
    return alt if os.path.exists(alt) else path


def compute_rows(sensex_csv, nifty_csv):
    """Run the per-second strategy on the live feed; return the display DataFrame
    (one row per trade) or an empty DataFrame if no trades yet."""
    df = load_series(sensex_csv, nifty_csv)          # per-second, tick_time/last_price
    trades = bt.backtest(df)                          # NO resample — per-second
    span = (df.date.min(), df.date.max(), len(df))
    if not len(trades):
        return pd.DataFrame(), span
    rows = []
    for _, r in trades.iterrows():
        LONG = (r.direction == "LONG_SPREAD")
        sxi, sxo, nfi, nfo = (("BUY", "SELL", "SELL", "BUY") if LONG
                              else ("SELL", "BUY", "BUY", "SELL"))
        if LONG:
            sp  = (r.sensex_exit - r.sensex_entry)
            npl = (r.nifty_entry - r.nifty_exit) * r.ratio
        else:
            sp  = (r.sensex_entry - r.sensex_exit)
            npl = (r.nifty_exit - r.nifty_entry) * r.ratio
        gross = (sp + npl) * PV
        net   = gross - CHARGE
        rows.append({
            "direction": "LONG" if LONG else "SHORT",
            "entry_time": pd.to_datetime(r.entry_date).strftime("%H:%M:%S"),
            "exit_time":  pd.to_datetime(r.exit_date).strftime("%H:%M:%S"),
            "SENSEX_in": sxi, "SENSEX_entry": round(float(r.sensex_entry), 1),
            "SENSEX_out": sxo, "SENSEX_exit": round(float(r.sensex_exit), 1),
            "NIFTY_in": nfi, "NIFTY_entry": round(float(r.nifty_entry), 1),
            "NIFTY_out": nfo, "NIFTY_exit": round(float(r.nifty_exit), 1),
            "entry_spread": round(float(r.spread_entry), 2),
            "exit_spread": round(float(r.spread_exit), 2),
            "entry_z": round(float(r.zscore_entry), 3),
            "exit_z": round(float(r.zscore_exit), 3),
            "pnl_pts": round(sp + npl, 2),
            "GROSS_Rs": round(gross), "charge_Rs": CHARGE, "NET_Rs": round(net),
            "P/L": "PROFIT" if net > 0 else "LOSS",
            "exit_reason": r.exit_reason,
        })
    d = pd.DataFrame(rows)
    d.insert(0, "#", range(1, len(d) + 1))
    d["running_NET_Rs"] = d.NET_Rs.cumsum()
    return d, span


def write_excel(day, d, fallback=True):
    """Write the 3-sheet report. With fallback=False, a locked file raises
    PermissionError (used by --watch so it retries instead of spawning _v2)."""
    gp = d.GROSS_Rs.sum(); nt = d.NET_Rs.sum()
    summ = pd.DataFrame({
        "metric": ["Date", "Source", "Basis", "Trades", "Wins", "Net win rate",
                   "GROSS", "Charge/trade", "NET (1 lot)", "Best", "Worst", "Updated"],
        "value": [day, "LIVE feed (sensex_today.csv / nifty_today.csv), per-second strategy",
                  "Rs20/order x4 = Rs80/round-trip", len(d),
                  int((d.NET_Rs > 0).sum()), f"{(d.NET_Rs > 0).mean()*100:.0f}%",
                  f"+Rs{gp:,.0f}/lot", f"Rs{CHARGE}",
                  f"Rs{nt:+,.0f}/lot ({'PROFIT' if nt > 0 else 'LOSS'})",
                  f"Rs{d.NET_Rs.max():+,.0f}", f"Rs{d.NET_Rs.min():+,.0f}",
                  datetime.now().strftime("%H:%M:%S")],
    })
    lots = pd.DataFrame([{
        "lots": L, "GROSS_Rs": round(gp * L), "brokerage_Rs(flat)": CHARGE * len(d),
        "NET_Rs": round(gp * L - CHARGE * len(d)), "margin_Rs": MARGIN_PER_LOT * L,
    } for L in LOTS])

    os.makedirs(OUT_DIR, exist_ok=True)
    out = os.path.join(OUT_DIR, f"live_trades_{day}.xlsx")
    try:
        xw = pd.ExcelWriter(out, engine="openpyxl")
    except PermissionError:
        if not fallback:
            raise
        out = out.replace(".xlsx", "_v2.xlsx"); xw = pd.ExcelWriter(out, engine="openpyxl")
    with xw as xl:
        summ.to_excel(xl, sheet_name="Summary", index=False)
        d.to_excel(xl, sheet_name="Live Trades (Rs20-order)", index=False)
        lots.to_excel(xl, sheet_name="Multiple Lots", index=False)
        for nm in xl.book.sheetnames:
            ws = xl.book[nm]
            for c in ws[1]:
                c.font = Font(bold=True, color="FFFFFF"); c.fill = HEAD
                c.alignment = Alignment(horizontal="center")
            ws.freeze_panes = "A2"
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 13
            hdr = {c.value: c.column for c in ws[1]}
            for col in ("GROSS_Rs", "NET_Rs", "running_NET_Rs"):
                ci = hdr.get(col)
                if not ci:
                    continue
                for rr in range(2, ws.max_row + 1):
                    cell = ws.cell(row=rr, column=ci)
                    try:
                        v = float(cell.value)
                    except (TypeError, ValueError):
                        continue
                    if v > 0:
                        cell.fill = GREEN
                    elif v < 0:
                        cell.fill = RED
            for col in ("SENSEX_in", "SENSEX_out", "NIFTY_in", "NIFTY_out", "P/L"):
                ci = hdr.get(col)
                if not ci:
                    continue
                for rr in range(2, ws.max_row + 1):
                    cell = ws.cell(row=rr, column=ci)
                    cell.fill = GREEN if cell.value in ("BUY", "PROFIT") else RED
    return out, gp, nt


def eod_reached():
    n = datetime.now()
    return (n.hour, n.minute) >= (EOD_HH, EOD_MM)


def main():
    day        = arg("--date", str(date.today()))
    sensex_csv = resolve(arg("--sensex", "sensex_today.csv"))
    nifty_csv  = resolve(arg("--nifty",  "nifty_today.csv"))
    watch      = "--watch" in sys.argv
    interval   = 15
    if watch:
        i = sys.argv.index("--watch")
        if i + 1 < len(sys.argv) and sys.argv[i + 1].isdigit():
            interval = int(sys.argv[i + 1])

    for p in (sensex_csv, nifty_csv):
        if not os.path.exists(p):
            print(f"Feed file not found: {p}\n"
                  f"Run from feed_data/ while tick_poller.py is/has been running.")
            return

    if not watch:                                    # ── one-shot ──
        d, span = compute_rows(sensex_csv, nifty_csv)
        print(f"Feed: {span[2]} ticks ({span[0]:%H:%M:%S} -> {span[1]:%H:%M:%S})")
        if not len(d):
            print(f"No trades from the feed yet (needs ~{bt.RATIO_LOOKBACK + bt.LOOKBACK} ticks warm-up).")
            return
        out, gp, nt = write_excel(day, d, fallback=True)
        print(f"Wrote {out}")
        print(f"{day}: {len(d)} live-feed trades | GROSS +Rs{gp:,.0f}/lot | "
              f"NET Rs{nt:+,.0f}/lot | win {(d.NET_Rs > 0).mean()*100:.0f}%")
        print(d[["#", "entry_time", "exit_time", "direction", "pnl_pts", "NET_Rs", "P/L"]]
              .tail(12).to_string(index=False))
        return

    # ── LIVE watch: rewrite the Excel on every NEW trade until 15:30 ──
    print(f"LIVE watch ON — rebuilding reports/live_trades_{day}.xlsx on every new "
          f"trade (poll {interval}s, until {EOD_HH:02d}:{EOD_MM:02d}). Keep the Excel CLOSED. Ctrl+C to stop.")
    last_written = -1
    try:
        while not eod_reached():
            try:
                d, _ = compute_rows(sensex_csv, nifty_csv)
                n = len(d)
                if n and n != last_written:
                    try:
                        out, gp, nt = write_excel(day, d, fallback=False)
                        last_written = n
                        print(f"  [{datetime.now():%H:%M:%S}] updated -> {n} trades | "
                              f"NET Rs{nt:+,.0f}/lot")
                    except PermissionError:
                        print(f"  [{datetime.now():%H:%M:%S}] {n} trades ready but Excel is OPEN "
                              f"- close it to update (will retry).")
            except Exception as e:
                print(f"  [{datetime.now():%H:%M:%S}] skip: {e}")
            time.sleep(interval)
    except KeyboardInterrupt:
        print("\nWatch stopped by user.")
    print(f"Watch ended ({datetime.now():%H:%M:%S}). Final file: reports/live_trades_{day}.xlsx")


if __name__ == "__main__":
    main()
