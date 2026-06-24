"""
multiday_report.py — combine several days of tick data and produce a multi-day
("long span") backtest report of the spread strategy:
  - each day backtested independently (no overnight-gap contamination), 1-min bars
  - aggregated totals, per-day breakdown, and LONG vs SHORT spread breakdown
  - a cross-day WALK-FORWARD: optimise on one day, test on the next (out-of-sample)

Outputs: multiday_per_day.csv, multiday_trades.csv, multiday_walkforward.csv,
         multiday_report.xlsx
"""
import pandas as pd
from openpyxl.styles import Font, PatternFill

import backtest as bt
from backtest_intraday import load_series, resample
from walkforward import optimize, evaluate   # grid optimise / out-of-sample evaluate

COST = 4.0    # Rs80 round-trip (Rs20/order x4) = 4 points
RULE = "1min"

DAYS = [
    ("2026-06-18", "sensex_Nifty/sensex_data.csv",          "sensex_Nifty/nifty_data.csv"),
    ("2026-06-19", "sensex_data_2026-06-19_archived.csv",   "nifty_data_2026-06-19_archived.csv"),
    ("2026-06-22", "sensex_data.csv",                       "nifty_data.csv"),
    ("2026-06-23", "sensex_today.csv",                      "nifty_today.csv"),
]


def fmt_headers(ws):
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E78")
    ws.freeze_panes = "A2"


def main():
    frames, per_day, all_trades = {}, [], []

    # ── per-day backtest (default strategy params) ───────────────────────────
    for d, sxf, nff in DAYS:
        try:
            df = resample(load_series(sxf, nff), RULE)
        except Exception as e:
            print(f"  {d}: SKIP ({e})")
            continue
        frames[d] = df
        tr = bt.backtest(df)
        if len(tr):
            tr = tr.copy()
            tr.insert(0, "day", d)
            tr["net_pnl_points"] = (tr["pnl_points"] - COST).round(2)
            all_trades.append(tr)
            L = tr[tr.direction == "LONG_SPREAD"]
            S = tr[tr.direction == "SHORT_SPREAD"]
            per_day.append({
                "day": d, "bars": len(df), "trades": len(tr),
                "gross_pts": round(tr.pnl_points.sum(), 2),
                "net_pts":   round(tr.net_pnl_points.sum(), 2),
                "long_trades": len(L),  "long_net":  round(L.pnl_points.sum() - COST*len(L), 2),
                "short_trades": len(S), "short_net": round(S.pnl_points.sum() - COST*len(S), 2),
            })
        else:
            per_day.append({"day": d, "bars": len(df), "trades": 0, "gross_pts": 0, "net_pts": 0,
                            "long_trades": 0, "long_net": 0, "short_trades": 0, "short_net": 0})

    pd_df = pd.DataFrame(per_day)
    trades = pd.concat(all_trades, ignore_index=True) if all_trades else pd.DataFrame()
    pd_df.to_csv("multiday_per_day.csv", index=False)
    if len(trades):
        trades.to_csv("multiday_trades.csv", index=False)

    # ── cross-day walk-forward ───────────────────────────────────────────────
    avail = list(frames.keys())
    wf = []
    for i in range(len(avail) - 1):
        tr_d, te_d = avail[i], avail[i + 1]
        cfg = optimize(frames[tr_d], COST, 3)
        if cfg is None:
            continue
        te = evaluate(frames[te_d], cfg, COST)
        wf.append({
            "train_day": tr_d, "test_day": te_d,
            "config": f"e{cfg['entry']}/x{cfg['exit']}/tp{cfg['tp']}/sl{cfg['sl']}/h{cfg['mh']}",
            "train_net_pts": cfg["net_pts"], "train_trades": cfg["trades"],
            "test_net_pts": (te or {}).get("net_pts"), "test_trades": (te or {}).get("trades", 0),
        })
    wf_df = pd.DataFrame(wf)
    if len(wf_df):
        wf_df.to_csv("multiday_walkforward.csv", index=False)

    # ── long vs short totals ─────────────────────────────────────────────────
    if len(trades):
        ls = trades.groupby("direction").agg(
            trades=("pnl_points", "size"),
            gross_pts=("pnl_points", "sum"),
            net_pts=("net_pnl_points", "sum"),
            wins=("net_pnl_points", lambda s: (s > 0).sum()),
        ).round(2).reset_index()
    else:
        ls = pd.DataFrame()

    # ── console summary ──────────────────────────────────────────────────────
    print("\n" + "=" * 74)
    print(f"  MULTI-DAY SPREAD REPORT  ({len(frames)} days, 1-min bars, cost {COST:.0f} pts/trade)")
    print("=" * 74)
    print(pd_df.to_string(index=False))
    print("-" * 74)
    tot_tr = int(pd_df.trades.sum()); tot_g = pd_df.gross_pts.sum(); tot_n = pd_df.net_pts.sum()
    print(f"  TOTAL: {tot_tr} trades | gross {tot_g:+.2f} | NET {tot_n:+.2f} pts")
    if len(ls):
        print("\n  LONG vs SHORT spread (net of cost):")
        print(ls.to_string(index=False))
    if len(wf_df):
        print("\n  CROSS-DAY WALK-FORWARD (best on train day -> next unseen day):")
        for _, r in wf_df.iterrows():
            tn = r.test_net_pts
            print(f"    {r.train_day} -> {r.test_day}:  train {r.train_net_pts:+.0f}  "
                  f"test(out) {tn:+.0f} pts" if tn is not None else
                  f"    {r.train_day} -> {r.test_day}:  train {r.train_net_pts:+.0f}  test: no trades")
        oos = wf_df.test_net_pts.dropna().sum()
        print(f"    TOTAL OUT-OF-SAMPLE: {oos:+.2f} pts")
    print("=" * 74)

    # ── Excel ────────────────────────────────────────────────────────────────
    with pd.ExcelWriter("multiday_report.xlsx", engine="openpyxl") as xl:
        pd_df.to_excel(xl, sheet_name="Per-Day Summary", index=False)
        if len(ls):     ls.to_excel(xl, sheet_name="Long vs Short", index=False)
        if len(wf_df):  wf_df.to_excel(xl, sheet_name="Walk-Forward", index=False)
        if len(trades): trades.to_excel(xl, sheet_name="All Trades", index=False)
        for ws in xl.book.worksheets:
            fmt_headers(ws)
    print("Wrote multiday_report.xlsx  (+ multiday_per_day.csv, multiday_trades.csv, multiday_walkforward.csv)")


if __name__ == "__main__":
    main()
