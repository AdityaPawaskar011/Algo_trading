r"""
oldfolder_report.py — backtest the INTRADAY SENSEX/NIFTY spread strategy across
ALL the historical 1-minute data in Old_data/ and write TWO Excel reports in the
SAME format as reports/this_week_trade_detail_v2.xlsx:

    old_data_monthwise.xlsx  — trades bucketed BY MONTH  (a report row per month)
    old_data_weekwise.xlsx   — trades bucketed BY WEEK   (a report row per week)

Strategy (identical to the v2 report — backtest.compute_signals + backtest):
    ratio  = rolling mean of Sensex/Nifty over 60 bars
    spread = Sensex - ratio*Nifty ;  z = (spread-mean)/std over 15 bars
    enter  |z| >= 2 ; exit on reversion |z| <= 0.7, stop -100, target +30,
    or after 15 bars (max-hold).
Run PER DAY on that day's 1-minute bars => purely intraday, no overnight carry.

Cost basis (the user's real basis): Rs20/order x 4 orders = Rs80/round-trip.
P&L = Rs20 per spread point per lot. Gross is rebuilt from the two legs using the
hedge ratio fixed at entry (1 SENSEX fut lot + 1 NIFTY fut lot ~ 3.2x hedge).

Old_data layout (written by old_data.py):
    Old_data/YYYY/MM_Month/ YYYY-MM-DD.csv          -> NIFTY  (use `close`)
                            sensex_YYYY-MM-DD.csv    -> SENSEX (use `close`)
    columns: timestamp,open,high,low,close,volume,open_interest

Run (from the code/ folder, with the project venv):
    ..\log_tradingVenv\Scripts\python.exe oldfolder_report.py
    python oldfolder_report.py --root Old_data --out-dir ..\reports
"""
import os
import sys
import glob

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment

import backtest as bt
from backtest_intraday import resample

# ── strategy params: EXACTLY the v2 intraday settings ──
bt.ENTRY, bt.EXIT, bt.STOP_LOSS, bt.PROFIT_TARGET, bt.MAX_HOLD, bt.RATIO_LOOKBACK, bt.LOOKBACK = \
    2.0, 0.7, 100, 30, 15, 60, 15

PV        = 20          # Rs per spread point per lot
PER_ORDER = 20          # Rs flat per order
CHARGE    = PER_ORDER * 4   # Rs80 round-trip (4 orders: 2 legs in + 2 legs out)
LOTS      = [1, 2, 5, 10, 20]
MARGIN_PER_LOT = 150000

GREEN = PatternFill("solid", fgColor="C6EFCE")
RED   = PatternFill("solid", fgColor="FFC7CE")
HEAD  = PatternFill("solid", fgColor="1F4E78")


def arg(flag, default):
    return sys.argv[sys.argv.index(flag) + 1] if flag in sys.argv else default


def discover(root):
    """Return {YYYY-MM-DD: (sensex_csv, nifty_csv)} for every matched day-pair."""
    days = {}
    for sx in glob.glob(os.path.join(root, "*", "*", "sensex_*.csv")):
        d = os.path.basename(sx)[len("sensex_"):-4]          # YYYY-MM-DD
        nf = os.path.join(os.path.dirname(sx), f"{d}.csv")    # plain = NIFTY
        if os.path.exists(nf):
            days[d] = (sx, nf)
    return dict(sorted(days.items()))


def load_day(sensex_csv, nifty_csv):
    """One day's 1-minute bars merged to date/sensex_close/nifty_close."""
    sx = (pd.read_csv(sensex_csv, usecols=["timestamp", "close"]).dropna()
            .drop_duplicates("timestamp").rename(columns={"close": "sensex_close"}))
    nf = (pd.read_csv(nifty_csv, usecols=["timestamp", "close"]).dropna()
            .drop_duplicates("timestamp").rename(columns={"close": "nifty_close"}))
    df = pd.merge(sx, nf, on="timestamp", how="inner")
    df["date"] = pd.to_datetime(df["timestamp"])
    df = df.sort_values("date").reset_index(drop=True)
    return df[["date", "sensex_close", "nifty_close"]]


def run_all(root):
    """Backtest every day intraday; return a DataFrame of per-trade rows."""
    days = discover(root)
    print(f"Found {len(days)} day-pairs in {root}/ "
          f"({min(days)} -> {max(days)})" if days else f"No day-pairs in {root}/")
    rows = []
    skipped = 0
    for date, (sx, nf) in days.items():
        df = load_day(sx, nf)
        df = resample(df, "1min")            # already 1-min; harmless de-dup
        if len(df) < bt.RATIO_LOOKBACK + bt.LOOKBACK + 2:
            skipped += 1
            continue
        t = bt.backtest(df)
        if not len(t):
            continue
        ts = pd.Timestamp(date)
        iso = ts.isocalendar()
        week_label = f"{iso[0]}-W{int(iso[1]):02d}"
        week_start = (ts.normalize() - pd.Timedelta(days=ts.weekday())).strftime("%Y-%m-%d")
        for _, r in t.iterrows():
            LONG = (r.direction == "LONG_SPREAD")
            et = pd.to_datetime(r.entry_date).strftime("%H:%M:%S")
            xt = pd.to_datetime(r.exit_date).strftime("%H:%M:%S")
            if LONG:
                sxi, sxo, nfi, nfo = "BUY", "SELL", "SELL", "BUY"
                sp  = (r.sensex_exit - r.sensex_entry)
                npl = (r.nifty_entry - r.nifty_exit) * r.ratio
            else:
                sxi, sxo, nfi, nfo = "SELL", "BUY", "BUY", "SELL"
                sp  = (r.sensex_entry - r.sensex_exit)
                npl = (r.nifty_exit - r.nifty_entry) * r.ratio
            gross = (sp + npl) * PV
            net   = gross - CHARGE
            rows.append({
                "date": date, "month": date[:7], "week": week_label, "week_start": week_start,
                "entry_time": et, "exit_time": xt, "trade": "LONG" if LONG else "SHORT",
                "SENSEX_in": sxi, "SENSEX_entry": round(r.sensex_entry, 1),
                "SENSEX_out": sxo, "SENSEX_exit": round(r.sensex_exit, 1),
                "NIFTY_in": nfi, "NIFTY_entry": round(r.nifty_entry, 1),
                "NIFTY_out": nfo, "NIFTY_exit": round(r.nifty_exit, 1),
                "exit_reason": r.exit_reason,
                "GROSS_Rs": round(gross), "charge_Rs": CHARGE, "NET_Rs": round(net),
                "P/L": "PROFIT" if net > 0 else "LOSS",
            })
    if skipped:
        print(f"  ({skipped} short/holiday days skipped)")
    d = pd.DataFrame(rows)
    if len(d):
        d.insert(0, "#", range(1, len(d) + 1))
        d["running_NET_Rs"] = d.NET_Rs.cumsum()
    return d


def bucket_summary(d, key, key_name):
    """One report row per bucket (month or week): per-lot NET (1/5/10 lots),
    chronological + running total. Brokerage is flat per trade, so it does NOT
    scale with lots — only the gross does."""
    g = d.groupby(key, sort=True)
    trades = g.size()
    gross  = g.GROSS_Rs.sum().round().astype(int)     # per 1 lot
    charge = trades * CHARGE                            # flat, same at any lot count
    s = pd.DataFrame({
        key_name: trades.index,
        "trades": trades.values,
        "GROSS_Rs": gross.values,
        "charge_Rs": charge.values,
        "NET_Rs": (gross - charge).values,             # 1 lot
        "NET_5lot": (gross * 5 - charge).values,
        "NET_10lot": (gross * 10 - charge).values,
        "win_%": (g.NET_Rs.apply(lambda s: (s > 0).mean() * 100)
                  .round().astype(int).values),
        "best_Rs": g.NET_Rs.max().round().astype(int).values,
        "worst_Rs": g.NET_Rs.min().round().astype(int).values,
    })
    s["running_NET_Rs"] = s.NET_Rs.cumsum()            # 1-lot cumulative
    return s


def lots_table(d):
    gp = d.GROSS_Rs.sum()
    return pd.DataFrame([{
        "lots": L, "GROSS_Rs": round(gp * L), "brokerage_Rs(flat)": CHARGE * len(d),
        "NET_Rs": round(gp * L - CHARGE * len(d)), "margin_Rs": MARGIN_PER_LOT * L,
    } for L in LOTS])


def overall_summary(d, period_label):
    gp = d.GROSS_Rs.sum(); nt = d.NET_Rs.sum()
    return pd.DataFrame({
        "metric": ["Period", "Strategy", "Basis", "Trades", "GROSS", "Charge/trade",
                   "NET (1 lot)", "Net win rate", "Best trade", "Worst trade"],
        "value": [period_label, "Intraday spread mean-reversion (per-day, 1-min bars)",
                  "Rs20/order x4 = Rs80/round-trip",
                  len(d), f"+Rs{gp:,.0f}/lot", f"Rs{CHARGE}",
                  f"Rs{nt:+,.0f}/lot ({'PROFIT' if nt > 0 else 'LOSS'})",
                  f"{(d.NET_Rs > 0).mean() * 100:.0f}%",
                  f"Rs{d.NET_Rs.max():+,.0f}", f"Rs{d.NET_Rs.min():+,.0f}"],
    })


def _style(xl, detail_sheet):
    for nm in xl.book.sheetnames:
        ws = xl.book[nm]
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF"); c.fill = HEAD
            c.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 13
    # colour money columns + decisions on every sheet that has them
    for nm in xl.book.sheetnames:
        ws = xl.book[nm]; hdr = {c.value: c.column for c in ws[1]}
        for col in ("GROSS_Rs", "NET_Rs", "NET_5lot", "NET_10lot",
                    "running_NET_Rs", "best_Rs", "worst_Rs"):
            ci = hdr.get(col)
            if not ci:
                continue
            for rr in range(2, ws.max_row + 1):
                cell = ws.cell(row=rr, column=ci)
                try:
                    v = float(cell.value)
                except (TypeError, ValueError):
                    continue
                if v > 0:
                    cell.fill = GREEN
                elif v < 0:
                    cell.fill = RED
        for col in ("SENSEX_in", "SENSEX_out", "NIFTY_in", "NIFTY_out", "P/L"):
            ci = hdr.get(col)
            if not ci:
                continue
            for rr in range(2, ws.max_row + 1):
                cell = ws.cell(row=rr, column=ci)
                cell.fill = GREEN if cell.value in ("BUY", "PROFIT") else RED


def write_report(d, key, key_name, breakdown_sheet, out_path):
    period = f"{d.date.min()} -> {d.date.max()}  ({len(d)} trades)"
    summ = overall_summary(d, period)
    brk  = bucket_summary(d, key, key_name)
    lots = lots_table(d)
    detail_cols = ["#", "date", "month", "week", "entry_time", "exit_time", "trade",
                   "SENSEX_in", "SENSEX_entry", "SENSEX_out", "SENSEX_exit",
                   "NIFTY_in", "NIFTY_entry", "NIFTY_out", "NIFTY_exit",
                   "exit_reason", "GROSS_Rs", "charge_Rs", "NET_Rs", "P/L", "running_NET_Rs"]
    try:
        xw = pd.ExcelWriter(out_path, engine="openpyxl")
    except Exception:
        out_path = out_path.replace(".xlsx", "_v2.xlsx")
        xw = pd.ExcelWriter(out_path, engine="openpyxl")
    with xw as xl:
        summ.to_excel(xl, sheet_name="Summary", index=False)
        brk.to_excel(xl, sheet_name=breakdown_sheet, index=False)
        det = d[detail_cols]
        for y in sorted(det["date"].str[:4].unique()):     # one Trade Detail sheet per year
            det[det["date"].str[:4] == y].to_excel(
                xl, sheet_name=f"Trades {y} (Rs20-order)", index=False)
        lots.to_excel(xl, sheet_name="Multiple Lots", index=False)
        _style(xl, "Trade Detail")
    print(f"Wrote {out_path}")
    return out_path, brk


def main():
    root    = arg("--root", os.path.join(os.path.dirname(os.path.abspath(__file__)), "Old_data"))
    out_dir = arg("--out-dir", os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "reports"))
    os.makedirs(out_dir, exist_ok=True)

    d = run_all(root)
    if not len(d):
        print("No trades generated across Old_data — nothing to write.")
        return

    gp = d.GROSS_Rs.sum(); nt = d.NET_Rs.sum()
    print("=" * 70)
    print(f"  OLD_DATA INTRADAY BACKTEST  ({d.date.min()} -> {d.date.max()})")
    print(f"  Basis Rs20/order = Rs80/round-trip | PV Rs20/point/lot")
    print("=" * 70)
    print(f"  Trades {len(d)} | GROSS +Rs{gp:,.0f}/lot | NET Rs{nt:+,.0f}/lot "
          f"({'PROFIT' if nt > 0 else 'LOSS'}) | net win {(d.NET_Rs > 0).mean()*100:.0f}%")
    print("=" * 70)

    m_path, m_brk = write_report(
        d, "month", "month", "Monthly Breakdown",
        os.path.join(out_dir, "old_data_monthwise.xlsx"))
    w_path, w_brk = write_report(
        d, "week", "week", "Weekly Breakdown",
        os.path.join(out_dir, "old_data_weekwise.xlsx"))

    pos_m = (m_brk.NET_Rs > 0).sum()
    print(f"\nMonths: {len(m_brk)} | net-positive {pos_m} | net-negative {len(m_brk) - pos_m}")
    print(m_brk[["month", "trades", "NET_Rs", "win_%", "running_NET_Rs"]].tail(12).to_string(index=False))


if __name__ == "__main__":
    main()
