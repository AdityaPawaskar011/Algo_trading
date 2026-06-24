"""
paper_trades_excel.py — Excel P&L report from paper_trades.csv, with rupee
profit/loss PER LOT.

Spread P&L is in index points. For a 1-lot pair (SENSEX fut lot 20 + NIFTY fut
lot ~65 ~= ratio*20), 1 spread point ~= Rs 20 per lot-pair, so
  rupee P&L per lot = points * 20      (override with --rs-per-point)
A realistic round-trip cost is ~35 pts (~Rs 700/lot), shown as a net view.

Run from feed_data/:
    python ..\\code\\paper_trades_excel.py --out ..\\reports\\paper_trades_pnl.xlsx
"""
import sys
import pandas as pd
from openpyxl.styles import Font, PatternFill

RS_PER_POINT = 20.0     # Rs per spread point per 1-lot pair (SENSEX fut lot size)
COST_PTS     = 35.0     # realistic round-trip cost in points

GREEN = PatternFill("solid", fgColor="C6EFCE")
RED   = PatternFill("solid", fgColor="FFC7CE")
HEAD  = PatternFill("solid", fgColor="1F4E78")


def arg(flag, default):
    return sys.argv[sys.argv.index(flag) + 1] if flag in sys.argv else default


def fmt(ws, money_cols=(), result_col=None):
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = HEAD
    ws.freeze_panes = "A2"
    hdr = {c.value: c.column for c in ws[1]}
    for nm in money_cols:
        ci = hdr.get(nm)
        if not ci:
            continue
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=ci)
            try:
                v = float(cell.value)
            except (TypeError, ValueError):
                continue
            if v > 0:   cell.fill = GREEN
            elif v < 0: cell.fill = RED
    if result_col and result_col in hdr:
        ci = hdr[result_col]
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=ci)
            cell.fill = GREEN if str(cell.value).upper() == "WIN" else RED


def main():
    src = arg("--in", "paper_trades.csv")
    rs  = float(arg("--rs-per-point", str(RS_PER_POINT)))
    out = arg("--out", "paper_trades_pnl.xlsx")

    t = pd.read_csv(src)
    if not len(t):
        print("No closed trades in", src, "yet.")
        return

    t["pnl_rs_per_lot"]   = (t["pnl_pts"] * rs).round(0)
    t["cum_pnl_pts"]      = t["pnl_pts"].cumsum().round(2)
    t["cum_rs_per_lot"]   = (t["cum_pnl_pts"] * rs).round(0)
    t["net_pts_after_cost"]      = (t["pnl_pts"] - COST_PTS).round(2)
    t["net_rs_per_lot_after_cost"] = (t["net_pts_after_cost"] * rs).round(0)
    t["result"] = t["pnl_pts"].apply(lambda p: "WIN" if p > 0 else "LOSS")

    cols = ["trade_date", "entry_time", "exit_time", "direction",
            "entry_zscore", "exit_zscore", "exit_reason",
            "pnl_pts", "pnl_rs_per_lot", "cum_pnl_pts", "cum_rs_per_lot",
            "net_pts_after_cost", "net_rs_per_lot_after_cost", "result"]
    trades = t[[c for c in cols if c in t.columns]]

    wins   = t[t.pnl_pts > 0]
    losses = t[t.pnl_pts < 0]
    gp, gl, net = wins.pnl_pts.sum(), losses.pnl_pts.sum(), t.pnl_pts.sum()
    net_cost = (t.pnl_pts - COST_PTS).sum()

    summary = pd.DataFrame({
        "metric": [
            "Total trades", "Wins", "Losses", "Win rate %",
            "-- GROSS (sandbox, no cost) --",
            "Gross profit (points)", "Gross profit (Rs/lot)",
            "Gross loss (points)", "Gross loss (Rs/lot)",
            "NET P&L (points)", "NET P&L (Rs/lot)",
            "Best trade (Rs/lot)", "Worst trade (Rs/lot)",
            f"-- AFTER ~{COST_PTS:.0f}pt cost (Rs{COST_PTS*rs:.0f}/lot per trade) --",
            "NET after cost (points)", "NET after cost (Rs/lot)",
            "-- conversion --",
            "Rs per point per lot", "Lot: SENSEX fut / NIFTY fut",
        ],
        "value": [
            len(t), len(wins), len(losses), round((t.pnl_pts > 0).mean()*100, 0),
            "",
            round(gp, 2), round(gp*rs, 0),
            round(gl, 2), round(gl*rs, 0),
            round(net, 2), round(net*rs, 0),
            round(wins.pnl_pts.max()*rs, 0) if len(wins) else 0,
            round(losses.pnl_pts.min()*rs, 0) if len(losses) else 0,
            "",
            round(net_cost, 2), round(net_cost*rs, 0),
            "",
            rs, "20 / 65",
        ],
    })

    ls_rows = []
    for d in ["LONG", "SHORT"]:
        s = t[t.direction == d]
        if len(s):
            ls_rows.append({"direction": d, "trades": len(s),
                            "pnl_points": round(s.pnl_pts.sum(), 2),
                            "pnl_Rs_per_lot": round(s.pnl_pts.sum()*rs, 0),
                            "win_pct": round((s.pnl_pts > 0).mean()*100, 0)})

    with pd.ExcelWriter(out, engine="openpyxl") as xl:
        summary.to_excel(xl, sheet_name="Summary", index=False)
        pd.DataFrame(ls_rows).to_excel(xl, sheet_name="Long vs Short", index=False)
        trades.to_excel(xl, sheet_name="Trades", index=False)
        fmt(xl.book["Summary"], money_cols=("value",))
        fmt(xl.book["Long vs Short"], money_cols=("pnl_points", "pnl_Rs_per_lot"))
        fmt(xl.book["Trades"],
            money_cols=("pnl_pts", "pnl_rs_per_lot", "cum_pnl_pts", "cum_rs_per_lot",
                        "net_pts_after_cost", "net_rs_per_lot_after_cost"),
            result_col="result")

    print(f"Wrote {out}")
    print(f"  {len(t)} trades | GROSS net {net:+.1f} pts = Rs {net*rs:+,.0f}/lot | "
          f"after cost: {net_cost:+.1f} pts = Rs {net_cost*rs:+,.0f}/lot")


if __name__ == "__main__":
    main()
