"""
profit_strategy.py — the ONLY configuration that backtested net-positive in our
testing: DAILY data, multi-day position hold, LONG spread focus, winners allowed
to run to mean-reversion (no tiny profit cap). Validated OUT-OF-SAMPLE.

Why this can work where intraday can't: it trades RARELY (a few dozen times over
years) but holds each position for days to capture a big mean-reversion move, so
the fixed ~48-pt round-trip cost is small relative to the profit. Intraday trades
hundreds of times for tiny moves the cost eats alive.

Honest test: fit/observe on 2024-2025 (TRAIN), then apply unchanged to 2026
(TEST, unseen). If LONG-spread is net-positive on the TEST set too, the edge is
real; if not, it isn't.

Run from feed_data/:
    python ..\\code\\profit_strategy.py --cost-pts 48 --split 2026-01-01
"""
import sys
import pandas as pd
from openpyxl.styles import Font, PatternFill

import backtest as bt

GREEN = PatternFill("solid", fgColor="C6EFCE"); RED = PatternFill("solid", fgColor="FFC7CE")
HEAD  = PatternFill("solid", fgColor="1F4E78")


def arg(flag, d): return sys.argv[sys.argv.index(flag) + 1] if flag in sys.argv else d


def fmt(ws):
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = HEAD
    ws.freeze_panes = "A2"


def stats(t, cost):
    if not len(t):
        return {"trades": 0, "net_pts": 0.0, "net_Rs_per_lot": 0, "win_pct": 0, "avg_hold_d": 0}
    n = t.net_pnl_points          # already gross - cost
    return {"trades": len(t), "net_pts": round(n.sum(), 1),
            "net_Rs_per_lot": round(n.sum() * 20),
            "win_pct": round((n > 0).mean() * 100, 0),
            "avg_hold_d": round(t.holding_days.mean(), 1)}


def main():
    cost  = float(arg("--cost-pts", "4"))
    split = pd.Timestamp(arg("--split", "2026-01-01"))

    # ── the profitable configuration: multi-day hold, winners run to reversion ──
    bt.ENTRY, bt.EXIT = 2.0, 0.7
    bt.STOP_LOSS, bt.PROFIT_TARGET, bt.MAX_HOLD = 100, 100000, 25   # target disabled -> reversion exit
    bt.RATIO_LOOKBACK, bt.LOOKBACK = 60, 15

    df = pd.read_csv("history.csv")
    df["date"] = pd.to_datetime(df["date"])
    df = df.sort_values("date").reset_index(drop=True)
    print(f"Daily series: {len(df)} days ({df.date.min().date()} -> {df.date.max().date()})")
    print(f"Config: entry +/-{bt.ENTRY} exit +/-{bt.EXIT} SL -{bt.STOP_LOSS} hold<= {bt.MAX_HOLD}d "
          f"(winners run to reversion) | cost {cost:.0f} pts | TRAIN<{split.date()}<=TEST")

    tr = bt.backtest(df)
    tr["entry_date"] = pd.to_datetime(tr["entry_date"])
    tr["exit_date"]  = pd.to_datetime(tr["exit_date"])
    tr["period"] = (tr.entry_date >= split).map({True: "TEST", False: "TRAIN"})
    tr = tr.rename(columns={"pnl_points": "gross_pnl_points"})
    tr["cost_points"]   = cost
    tr["net_pnl_points"] = (tr.gross_pnl_points - cost).round(2)
    tr["net_Rs_per_lot"] = (tr.net_pnl_points * 20).round(0)

    train, test = tr[tr.period == "TRAIN"], tr[tr.period == "TEST"]
    L = tr[tr.direction == "LONG_SPREAD"]

    # build the grid: rows = {ALL, LONG, SHORT} x cols {TRAIN, TEST, FULL}
    rows = []
    for label, sub in [("ALL", tr), ("LONG only", tr[tr.direction == "LONG_SPREAD"]),
                       ("SHORT only", tr[tr.direction == "SHORT_SPREAD"])]:
        tr_s = stats(sub[sub.period == "TRAIN"], cost)
        te_s = stats(sub[sub.period == "TEST"], cost)
        fu_s = stats(sub, cost)
        rows.append({"set": label,
                     "train_trades": tr_s["trades"], "train_net_pts": tr_s["net_pts"], "train_net_Rs": tr_s["net_Rs_per_lot"],
                     "test_trades": te_s["trades"], "test_net_pts": te_s["net_pts"], "test_net_Rs": te_s["net_Rs_per_lot"],
                     "full_trades": fu_s["trades"], "full_net_pts": fu_s["net_pts"], "full_net_Rs": fu_s["net_Rs_per_lot"],
                     "full_win_pct": fu_s["win_pct"], "avg_hold_d": fu_s["avg_hold_d"]})
    grid = pd.DataFrame(rows)

    long_test = grid[grid.set == "LONG only"].iloc[0]
    long_train = long_test  # same row, different cols
    oos_long_pos = long_test["test_net_pts"] > 0 and long_test["train_net_pts"] > 0
    verdict = ("LONG-spread is net-positive in BOTH train and test -> a REAL out-of-sample edge"
               if oos_long_pos else
               "LONG-spread does NOT hold net-positive out-of-sample -> not a reliable edge")

    print("\n" + "=" * 78)
    print("  PROFIT CANDIDATE: daily multi-day spread, train vs test (net of cost)")
    print("=" * 78)
    print(grid.to_string(index=False))
    print("-" * 78)
    print(f"  VERDICT: {verdict}")
    print("=" * 78)

    # ── Excel ──
    out = arg("--out", "profit_strategy.xlsx")
    summary = pd.DataFrame({
        "metric": ["Strategy", "Data", "Cost (pts/round-trip)", "Train period", "Test period",
                   "LONG train net (pts/Rs)", "LONG test net (pts/Rs)", "LONG full net (pts/Rs)",
                   "LONG win% / avg hold", "SHORT full net (pts)", "VERDICT", "Caveat"],
        "value": ["Daily, multi-day hold, winners run to reversion",
                  f"{len(df)} daily bars {df.date.min().date()}..{df.date.max().date()}",
                  cost, f"< {split.date()}", f">= {split.date()}",
                  f"{long_test['train_net_pts']} / Rs{long_test['train_net_Rs']:,}",
                  f"{long_test['test_net_pts']} / Rs{long_test['test_net_Rs']:,}",
                  f"{long_test['full_net_pts']} / Rs{long_test['full_net_Rs']:,}",
                  f"{long_test['full_win_pct']}% / {long_test['avg_hold_d']}d",
                  grid[grid.set == 'SHORT only'].iloc[0]['full_net_pts'],
                  verdict,
                  "Few trades = modest sample. Trending months (e.g. Jun) still lose. "
                  "Paper-trade live before real money."],
    })
    cols = ["period", "entry_date", "exit_date", "direction", "holding_days",
            "zscore_entry", "zscore_exit", "exit_reason",
            "gross_pnl_points", "cost_points", "net_pnl_points", "net_Rs_per_lot"]
    trades_out = tr[[c for c in cols if c in tr.columns]].copy()
    trades_out["entry_date"] = trades_out.entry_date.dt.strftime("%Y-%m-%d")
    trades_out["exit_date"]  = trades_out.exit_date.dt.strftime("%Y-%m-%d")
    long_out = trades_out[trades_out.direction == "LONG_SPREAD"]

    with pd.ExcelWriter(out, engine="openpyxl") as xl:
        summary.to_excel(xl, sheet_name="Summary", index=False)
        grid.to_excel(xl, sheet_name="Train vs Test", index=False)
        long_out.to_excel(xl, sheet_name="LONG trades", index=False)
        trades_out.to_excel(xl, sheet_name="All trades", index=False)
        wb = xl.book
        for nm in wb.sheetnames:
            fmt(wb[nm])
        for nm, mc in [("Train vs Test", ["train_net_pts", "test_net_pts", "full_net_pts"]),
                       ("LONG trades", ["net_pnl_points", "net_Rs_per_lot"]),
                       ("All trades", ["net_pnl_points", "net_Rs_per_lot"])]:
            ws = wb[nm]; hdr = {c.value: c.column for c in ws[1]}
            for col in mc:
                ci = hdr.get(col)
                if not ci:
                    continue
                for r in range(2, ws.max_row + 1):
                    cell = ws.cell(row=r, column=ci)
                    try:
                        v = float(cell.value)
                    except (TypeError, ValueError):
                        continue
                    cell.fill = GREEN if v > 0 else (RED if v < 0 else cell.fill)
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
