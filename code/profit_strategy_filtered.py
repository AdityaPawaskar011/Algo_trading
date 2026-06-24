"""
profit_strategy_filtered.py — daily multi-day spread strategy + a TREND FILTER,
validated out-of-sample (train 2024-25, test 2026).

Trend filter (the one honest improvement worth trying): don't FADE a strong
trend. Measure the spread's move over the last TREND_WIN days, normalised by its
std (trend_norm). Skip a mean-reversion entry when the trend is strongly against
it:
  - SHORT signal (z>=+entry) but spread trending UP  (trend_norm > +THR) -> skip
  - LONG  signal (z<=-entry) but spread trending DOWN (trend_norm < -THR) -> skip
This is exactly the June failure mode (shorting a spread that kept rising).

Filter params are chosen a-priori (TREND_WIN=20, THR=1.0 std) and NOT tuned to
the test set. Runs filter ON and OFF for honest comparison.

Run from feed_data/:
    python ..\\code\\profit_strategy_filtered.py --cost-pts 48 --split 2026-01-01
"""
import sys
import numpy as np
import pandas as pd
from openpyxl.styles import Font, PatternFill

# a-priori params (not tuned to test)
RATIO_WIN, SPREAD_WIN, TREND_WIN = 60, 15, 20
ENTRY, EXIT, STOP, MAXHOLD = 2.0, 0.7, 100, 25
TREND_THR = 1.0
PV = 20  # Rs per spread point per lot

GREEN = PatternFill("solid", fgColor="C6EFCE"); RED = PatternFill("solid", fgColor="FFC7CE")
HEAD = PatternFill("solid", fgColor="1F4E78")


def arg(flag, d): return sys.argv[sys.argv.index(flag) + 1] if flag in sys.argv else d


def signals(df):
    d = df.copy()
    d["ratio"]  = (d.sensex_close / d.nifty_close).rolling(RATIO_WIN).mean()
    d["spread"] = d.sensex_close - d.ratio * d.nifty_close
    d["sma"]    = d.spread.rolling(SPREAD_WIN).mean()
    d["ssd"]    = d.spread.rolling(SPREAD_WIN).std()
    d["z"]      = (d.spread - d.sma) / d.ssd
    d["trend_norm"] = (d.spread - d.spread.shift(TREND_WIN)) / d.ssd
    return d


def backtest(df, cost, use_filter):
    d = signals(df).reset_index(drop=True)
    n = len(d); start = max(RATIO_WIN, SPREAD_WIN, TREND_WIN); pos = None; out = []
    for i in range(start, n):
        r = d.iloc[i]; z, tn = r.z, r.trend_norm
        if np.isnan(z) or np.isnan(tn):
            continue
        if pos is not None:
            held = i - pos["i"]
            live = (r.spread - pos["spread"]) if pos["dir"] == "LONG" else (pos["spread"] - r.spread)
            why = ""
            if   live <= -STOP:                              why = "stop_loss"
            elif pos["dir"] == "LONG"  and z >= -EXIT:       why = "reverted"
            elif pos["dir"] == "SHORT" and z <=  EXIT:       why = "reverted"
            elif held >= MAXHOLD:                            why = "max_hold"
            if why:
                out.append({"entry_date": pos["date"], "exit_date": r.date, "direction": pos["dir"],
                            "holding_days": held, "zscore_entry": round(pos["z"], 2),
                            "trend_at_entry": round(pos["tn"], 2), "zscore_exit": round(z, 2),
                            "exit_reason": why, "gross_pnl_points": round(live, 2)})
                pos = None
        if pos is None:
            if z <= -ENTRY and not (use_filter and tn < -TREND_THR):
                pos = {"dir": "LONG", "i": i, "date": r.date, "spread": r.spread, "z": z, "tn": tn}
            elif z >= ENTRY and not (use_filter and tn > TREND_THR):
                pos = {"dir": "SHORT", "i": i, "date": r.date, "spread": r.spread, "z": z, "tn": tn}
    t = pd.DataFrame(out)
    if len(t):
        t["cost_points"] = cost
        t["net_pnl_points"] = (t.gross_pnl_points - cost).round(2)
        t["net_Rs_per_lot"] = (t.net_pnl_points * PV).round(0)
    return t


def net(t):
    return round(t.net_pnl_points.sum(), 1) if len(t) else 0.0


def main():
    cost  = float(arg("--cost-pts", "4"))
    split = pd.Timestamp(arg("--split", "2026-01-01"))
    out   = arg("--out", "profit_strategy_filtered.xlsx")

    df = pd.read_csv("history.csv")
    df["date"] = pd.to_datetime(df["date"])
    df = df.sort_values("date").reset_index(drop=True)
    print(f"Daily series: {len(df)} days ({df.date.min().date()} -> {df.date.max().date()}) | cost {cost:.0f} pts")
    print(f"Filter: skip fading a trend > {TREND_THR} std over {TREND_WIN}d | entry +/-{ENTRY} exit +/-{EXIT}")

    rows = []
    results = {}
    for use_filter in (False, True):
        t = backtest(df, cost, use_filter)
        t["entry_date"] = pd.to_datetime(t["entry_date"])
        results[use_filter] = t
        for label, sub in [("ALL", t), ("LONG", t[t.direction == "LONG"])]:
            tr = sub[sub.entry_date < split]; te = sub[sub.entry_date >= split]
            rows.append({"filter": "ON" if use_filter else "OFF", "set": label,
                         "train_trades": len(tr), "train_net_pts": net(tr), "train_Rs": round(net(tr)*PV),
                         "test_trades": len(te), "test_net_pts": net(te), "test_Rs": round(net(te)*PV),
                         "full_net_pts": net(sub), "full_Rs": round(net(sub)*PV)})
    grid = pd.DataFrame(rows)

    # compare LONG out-of-sample: filter OFF vs ON
    long_off = grid[(grid["filter"] == "OFF") & (grid["set"] == "LONG")].iloc[0]
    long_on  = grid[(grid["filter"] == "ON") & (grid["set"] == "LONG")].iloc[0]
    improved = long_on.test_net_pts > long_off.test_net_pts
    oos_pos  = long_on.test_net_pts > 0
    verdict = (f"FILTER HELPS + LONG net-positive out-of-sample (test {long_on.test_net_pts:+.0f} pts / "
               f"Rs{long_on.test_Rs:+,.0f}) -> worth paper-trading"
               if oos_pos else
               f"Filter {'improves' if improved else 'does not fix'} it, but LONG is STILL not net-positive "
               f"out-of-sample (test {long_on.test_net_pts:+.0f} pts) -> not a reliable edge")

    print("\n" + "=" * 90)
    print("  TREND-FILTER TEST: net of cost, train (2024-25) vs test (2026, unseen)")
    print("=" * 90)
    print(grid.to_string(index=False))
    print("-" * 90)
    print(f"  LONG out-of-sample:  filter OFF {long_off.test_net_pts:+.0f} pts  ->  filter ON {long_on.test_net_pts:+.0f} pts")
    print(f"  VERDICT: {verdict}")
    print("=" * 90)

    # ── Excel ──
    summary = pd.DataFrame({
        "metric": ["Strategy", "Trend filter", "Cost (pts)", "Split",
                   "LONG test net OFF (pts/Rs)", "LONG test net ON (pts/Rs)",
                   "Filter helps OOS?", "VERDICT", "Caveat"],
        "value": ["Daily multi-day spread, winners run to reversion",
                  f"skip fading trend > {TREND_THR} std / {TREND_WIN}d", cost, str(split.date()),
                  f"{long_off.test_net_pts} / Rs{long_off.test_Rs:,}",
                  f"{long_on.test_net_pts} / Rs{long_on.test_Rs:,}",
                  "YES" if improved else "no",
                  verdict,
                  "Small trade sample; one market history. Validate further before real money."],
    })
    ton = results[True].copy()
    if len(ton):
        ton["entry_date"] = ton.entry_date.dt.strftime("%Y-%m-%d")
        ton["exit_date"]  = pd.to_datetime(ton.exit_date).dt.strftime("%Y-%m-%d")
    with pd.ExcelWriter(out, engine="openpyxl") as xl:
        summary.to_excel(xl, sheet_name="Summary", index=False)
        grid.to_excel(xl, sheet_name="Filter ON vs OFF", index=False)
        (ton if len(ton) else pd.DataFrame()).to_excel(xl, sheet_name="Trades (filter ON)", index=False)
        wb = xl.book
        for nm in wb.sheetnames:
            ws = wb[nm]
            for c in ws[1]:
                c.font = Font(bold=True, color="FFFFFF"); c.fill = HEAD
            ws.freeze_panes = "A2"
        ws = wb["Filter ON vs OFF"]; hdr = {c.value: c.column for c in ws[1]}
        for col in ("train_net_pts", "test_net_pts", "full_net_pts"):
            ci = hdr.get(col)
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=ci)
                try: v = float(cell.value)
                except (TypeError, ValueError): continue
                cell.fill = GREEN if v > 0 else (RED if v < 0 else cell.fill)
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
