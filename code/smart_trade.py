r"""
smart_trade.py — "smart" intraday backtest that applies the EDGE logic (the same
quality-over-quantity filter now in the live paper trader) to THIS WEEK's
per-second feed, and writes a detailed Excel: reports/smart_trade.xlsx.

EDGE logic (per day, intraday, no overnight carry):
    spread = Sensex - ratio*Nifty (ratio = 60-tick mean); z over 15 ticks.
    ENTER  |z| >= 2  AND  room-to-revert (|spread - mean|) >= TARGET
    EXIT   +TARGET (book) / -STOP / MAXHOLD ticks / EOD square-off
This only takes setups whose reversion is several times the Rs80 charge, so it
does NOT over-trade on shallow ~4-pt wiggles (which is what bled the old logic).

Cost basis: Rs20/order x4 = Rs80/round-trip. P&L = Rs20 per spread point per lot.

Run from anywhere (paths are absolute below) with the project venv:
    ..\log_tradingVenv\Scripts\python.exe ..\code\smart_trade.py --target 8
    python ..\code\smart_trade.py --target 8 --out ..\reports\smart_trade.xlsx
"""
import os
import sys
import numpy as np
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from backtest_intraday import load_series

RATIO, SPRWIN, ENTRY = 60, 15, 2.0
STOP, MAXHOLD = 100, 1800           # stop (pts), max hold (ticks ~ seconds)
PV = 20                              # Rs per spread point per lot
COST_PTS = 4                         # Rs80 round-trip / Rs20 = 4 pts
CHARGE = COST_PTS * PV               # Rs80
LOTS = [1, 2, 5, 10, 20]
MARGIN_PER_LOT = 150000

FD = r"c:\Users\ADITYA\Desktop\Algo_trading\feed_data"
# This week's per-second day-pairs (file naming differs by day; only existing ones used)
WEEK = [
    ("2026-06-22", FD + r"\sensex_data.csv",                                 FD + r"\nifty_data.csv"),
    ("2026-06-23", FD + r"\2026_06_23\sensex_today_2026-06-23_archived.csv", FD + r"\2026_06_23\nifty_today_2026-06-23_archived.csv"),
    ("2026-06-24", FD + r"\2026_06_24\sensex_today_2026-06-24_archived.csv", FD + r"\2026_06_24\nifty_today_2026-06-24_archived.csv"),
    ("2026-06-25", FD + r"\sensex_today.csv",                                FD + r"\nifty_today.csv"),
]

GREEN = PatternFill("solid", fgColor="C6EFCE")
RED   = PatternFill("solid", fgColor="FFC7CE")
HEAD  = PatternFill("solid", fgColor="1F4E78")


def arg(flag, d): return sys.argv[sys.argv.index(flag) + 1] if flag in sys.argv else d


def sig(df):
    d = df.copy()
    d["ratio"]  = (d.sensex_close / d.nifty_close).rolling(RATIO).mean()
    d["spread"] = d.sensex_close - d.ratio * d.nifty_close
    d["sma"]    = d.spread.rolling(SPRWIN).mean()
    d["ssd"]    = d.spread.rolling(SPRWIN).std()
    d["z"]      = (d.spread - d.sma) / d.ssd
    return d


def bt_edge(s, target, stop=STOP, maxhold=MAXHOLD):
    """Return a list of trade dicts (entry idx ei, exit idx xi, dir, gross_pts, reason)."""
    n = len(s); start = max(RATIO, SPRWIN); pos = None; out = []
    for i in range(start, n):
        r = s.iloc[i]; z = r.z
        if np.isnan(z):
            continue
        if pos is not None:
            held = i - pos["ei"]
            live = (r.spread - pos["spread"]) if pos["dir"] == "LONG" else (pos["spread"] - r.spread)
            why = ("target" if live >= target else "stop" if live <= -stop else
                   "maxhold" if held >= maxhold else "")
            if why:
                out.append({**pos, "xi": i, "gross_pts": live, "reason": why}); pos = None
        if pos is None:
            room = abs(r.spread - r.sma)
            if z <= -ENTRY and room >= target:
                pos = {"dir": "LONG", "ei": i, "spread": r.spread, "z": z}
            elif z >= ENTRY and room >= target:
                pos = {"dir": "SHORT", "ei": i, "spread": r.spread, "z": z}
    if pos is not None:
        r = s.iloc[-1]
        live = (r.spread - pos["spread"]) if pos["dir"] == "LONG" else (pos["spread"] - r.spread)
        out.append({**pos, "xi": n - 1, "gross_pts": live, "reason": "EOD"})
    return out


def day_rows(date, s, trades):
    rows = []
    for t in trades:
        e = s.iloc[t["ei"]]; x = s.iloc[t["xi"]]
        LONG = (t["dir"] == "LONG")
        gp = t["gross_pts"]
        gross = round(gp * PV); net = round((gp - COST_PTS) * PV)
        rows.append({
            "date": date,
            "entry_time": e.date.strftime("%H:%M:%S"), "exit_time": x.date.strftime("%H:%M:%S"),
            "hold_sec": int((x.date - e.date).total_seconds()),
            "trade": t["dir"],
            "SENSEX_in": "BUY" if LONG else "SELL", "SENSEX_entry": round(float(e.sensex_close), 1),
            "SENSEX_out": "SELL" if LONG else "BUY", "SENSEX_exit": round(float(x.sensex_close), 1),
            "NIFTY_in": "SELL" if LONG else "BUY", "NIFTY_entry": round(float(e.nifty_close), 1),
            "NIFTY_out": "BUY" if LONG else "SELL", "NIFTY_exit": round(float(x.nifty_close), 1),
            "entry_z": round(float(t["z"]), 2), "exit_reason": t["reason"],
            "GROSS_Rs": gross, "charge_Rs": CHARGE, "NET_Rs": net,
            "P/L": "PROFIT" if net > 0 else "LOSS",
        })
    return rows


def style(xl):
    for nm in xl.book.sheetnames:
        ws = xl.book[nm]
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF"); c.fill = HEAD
            c.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 12
        hdr = {c.value: c.column for c in ws[1]}
        for col in ("GROSS_Rs", "NET_Rs", "running_NET_Rs", "NET_5lot", "NET_10lot"):
            ci = hdr.get(col)
            if not ci:
                continue
            for rr in range(2, ws.max_row + 1):
                cell = ws.cell(row=rr, column=ci)
                try:
                    v = float(cell.value)
                except (TypeError, ValueError):
                    continue
                if v > 0:
                    cell.fill = GREEN
                elif v < 0:
                    cell.fill = RED
        for col in ("SENSEX_in", "SENSEX_out", "NIFTY_in", "NIFTY_out", "P/L"):
            ci = hdr.get(col)
            if not ci:
                continue
            for rr in range(2, ws.max_row + 1):
                cell = ws.cell(row=rr, column=ci)
                cell.fill = GREEN if cell.value in ("BUY", "PROFIT") else RED


def main():
    target = float(arg("--target", "8"))
    out    = arg("--out", r"c:\Users\ADITYA\Desktop\Algo_trading\reports\smart_trade.xlsx")
    print(f"SMART intraday EDGE backtest | target {target:.0f} pts | charge Rs{CHARGE} | this week\n")

    all_rows, per_day = [], []
    for date, sx, nf in WEEK:
        if not (os.path.exists(sx) and os.path.exists(nf)):
            print(f"  MISSING {date} - skipped"); continue
        s = sig(load_series(sx, nf)).reset_index(drop=True)
        trades = bt_edge(s, target)
        rows = day_rows(date, s, trades)
        all_rows += rows
        net = sum(r["NET_Rs"] for r in rows)
        wins = sum(1 for r in rows if r["NET_Rs"] > 0)
        per_day.append({"date": date, "ticks": len(s), "trades": len(rows),
                        "NET_Rs": net, "NET_5lot": net + 4 * len(rows) * PV * 0,  # placeholder fixed below
                        "win_%": round(wins / len(rows) * 100) if rows else 0})
        print(f"  {date}: {len(s):>6} ticks -> {len(rows):>4} trades | NET Rs{net:>+10,.0f} | "
              f"win {round(wins/len(rows)*100) if rows else 0}%")

    if not all_rows:
        print("No trades / no data."); return

    d = pd.DataFrame(all_rows)
    d.insert(0, "#", range(1, len(d) + 1))
    d["running_NET_Rs"] = d.NET_Rs.cumsum()

    # per-day breakdown (gross/charge/net + multi-lot; brokerage flat -> doesn't scale)
    g = d.groupby("date", sort=True)
    gross = g.GROSS_Rs.sum(); ntr = g.size(); chg = ntr * CHARGE
    pdb = pd.DataFrame({
        "date": gross.index, "trades": ntr.values,
        "GROSS_Rs": gross.values.round().astype(int), "charge_Rs": chg.values,
        "NET_Rs": (gross.values - chg.values).round().astype(int),
        "NET_5lot": (gross.values * 5 - chg.values).round().astype(int),
        "NET_10lot": (gross.values * 10 - chg.values).round().astype(int),
        "win_%": g.NET_Rs.apply(lambda s: round((s > 0).mean() * 100)).values,
    })
    pdb["running_NET_Rs"] = pdb.NET_Rs.cumsum()

    GP = d.GROSS_Rs.sum(); NT = d.NET_Rs.sum(); n = len(d)
    lots = pd.DataFrame([{
        "lots": L, "GROSS_Rs": round(GP * L), "brokerage_Rs(flat)": CHARGE * n,
        "NET_Rs": round(GP * L - CHARGE * n), "margin_Rs": MARGIN_PER_LOT * L,
    } for L in LOTS])

    summary = pd.DataFrame({
        "metric": ["Strategy", "Basis", "Entry filter", "Exit", "Target", "Period",
                   "Days", "Trades", "GROSS", "NET (1 lot)", "Net win rate", "Best", "Worst", "Caveat"],
        "value": ["Intraday EDGE spread mean-reversion (per-day, per-second ticks)",
                  "Rs20/order x4 = Rs80/round-trip",
                  f"|z| >= {ENTRY} AND room-to-revert >= {target:.0f} pts",
                  f"+{target:.0f} / -{STOP} / {MAXHOLD} ticks / EOD",
                  f"{target:.0f} pts (Rs{target*PV:.0f})",
                  f"{d.date.min()} -> {d.date.max()}", d.date.nunique(), n,
                  f"+Rs{GP:,.0f}/lot", f"Rs{NT:+,.0f}/lot ({'PROFIT' if NT > 0 else 'LOSS'})",
                  f"{(d.NET_Rs > 0).mean()*100:.0f}%", f"Rs{d.NET_Rs.max():+,.0f}", f"Rs{d.NET_Rs.min():+,.0f}",
                  "Per-second CLOSE fills, no slippage/bid-ask modeled; calm week. In-sample; "
                  "validate forward (paper) before real money."]})

    detail_cols = ["#", "date", "entry_time", "exit_time", "hold_sec", "trade",
                   "SENSEX_in", "SENSEX_entry", "SENSEX_out", "SENSEX_exit",
                   "NIFTY_in", "NIFTY_entry", "NIFTY_out", "NIFTY_exit",
                   "entry_z", "exit_reason", "GROSS_Rs", "charge_Rs", "NET_Rs", "P/L", "running_NET_Rs"]
    try:
        xw = pd.ExcelWriter(out, engine="openpyxl")
    except Exception:
        out = out.replace(".xlsx", "_v2.xlsx"); xw = pd.ExcelWriter(out, engine="openpyxl")
    with xw as xl:
        summary.to_excel(xl, sheet_name="Summary", index=False)
        pdb.to_excel(xl, sheet_name="Per-Day", index=False)
        d[detail_cols].to_excel(xl, sheet_name="Trade Detail (Rs20-order)", index=False)
        lots.to_excel(xl, sheet_name="Multiple Lots", index=False)
        style(xl)
    print(f"\nTOTAL: {n} trades | GROSS +Rs{GP:,.0f}/lot | NET Rs{NT:+,.0f}/lot | "
          f"net win {(d.NET_Rs > 0).mean()*100:.0f}%")
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
