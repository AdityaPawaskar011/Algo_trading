"""
Generate a fully formatted Excel report of the backtest.

Usage:
    python report.py                              # yfinance, 2023-01-01 to today
    python report.py --source upstox             # use upstox_feed
    python report.py --start 2023-01-01          # custom start date
    python report.py --start 2023-01-01 --end 2025-12-31
    python report.py --out my_report.xlsx        # custom output filename

Sheets produced:
    1. Summary          - key metrics + strategy parameters
    2. Trade Log        - every trade with entry/exit details (green=win, red=loss)
    3. Day-wise Log     - every trading day with price, ratio, spread, z-score, status
    4. Monthly P&L      - month-by-month breakdown
"""

import sys
import numpy as np
import pandas as pd
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from backtest import (
    load_from_sql, compute_signals, backtest, summary,
    ENTRY, EXIT, STOP_LOSS, MAX_HOLD, RATIO_LOOKBACK, LOOKBACK, MODE,
)

# ── Colour palette ─────────────────────────────────────────────────────────────
C_NAVY     = "1F4E79"
C_BLUE     = "2E75B6"
C_LT_BLUE  = "BDD7EE"
C_WIN_BG   = "C6EFCE"; C_WIN_FG  = "276221"
C_LOSS_BG  = "FFC7CE"; C_LOSS_FG = "9C0006"
C_SL_BG    = "FFEB9C"; C_SL_FG   = "9C6500"
C_EXP_BG   = "FCE4D6"; C_EXP_FG  = "843C0C"
C_WHITE    = "FFFFFF"
C_LGRAY    = "F5F5F5"
C_GRAY     = "D9D9D9"
C_YELLOW   = "FFF2CC"

thin = Side(style="thin", color="AAAAAA")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def fill(c): return PatternFill("solid", fgColor=c)
def font(c=None, bold=False, sz=10): return Font(color=c, bold=bold, size=sz)
def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)


def hcell(ws, r, c, val, bg=C_NAVY, fg=C_WHITE, bold=True, sz=10):
    cell = ws.cell(row=r, column=c, value=val)
    cell.fill = fill(bg)
    cell.font = font(fg, bold=bold, sz=sz)
    cell.alignment = center()
    cell.border = BORDER
    return cell


def dcell(ws, r, c, val, bg=None, fg=None, bold=False, align="center", fmt=None):
    cell = ws.cell(row=r, column=c, value=val)
    if bg: cell.fill = fill(bg)
    if fg or bold: cell.font = font(fg, bold=bold)
    cell.alignment = center() if align == "center" else left()
    cell.border = BORDER
    if fmt: cell.number_format = fmt
    return cell


def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


# ── Load & compute ─────────────────────────────────────────────────────────────

def load_data(source, start_date, end_date):
    df = load_from_sql(source)
    if start_date:
        df = df[df["date"] >= pd.Timestamp(start_date)]
    if end_date:
        df = df[df["date"] <= pd.Timestamp(end_date)]
    df = df.reset_index(drop=True)
    signals = compute_signals(df)
    trades  = backtest(df)
    return signals, trades


# ── Sheet 1: Summary ──────────────────────────────────────────────────────────

def write_summary(ws, trades, signals, source, start_date, end_date):
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "SENSEX - NIFTY RATIO SPREAD STRATEGY  |  BACKTEST REPORT"
    t.fill  = fill(C_NAVY)
    t.font  = font(C_WHITE, bold=True, sz=14)
    t.alignment = center()
    ws.row_dimensions[1].height = 30

    # Sub-title
    ws.merge_cells("A2:H2")
    d1 = str(signals["date"].min().date())
    d2 = str(signals["date"].max().date())
    sub = ws["A2"]
    sub.value = f"Data Source: {source.upper()}   |   Period: {d1}  to  {d2}   |   Generated: {date.today()}"
    sub.fill  = fill(C_BLUE)
    sub.font  = font(C_WHITE, sz=10)
    sub.alignment = center()

    # ── Performance metrics ──
    wins     = trades[trades.pnl_points > 0]
    losses   = trades[trades.pnl_points <= 0]
    stopped  = trades[trades.exit_reason == "stop_loss"]
    expired  = trades[trades.exit_reason == "expiry"]
    reverted = trades[trades.exit_reason == "reverted"]

    metrics = [
        ("PERFORMANCE METRICS", "", "STRATEGY PARAMETERS", ""),
        ("Total Trades",       len(trades),           "Entry Z-Score Threshold", f"+/- {ENTRY}"),
        ("Winning Trades",     len(wins),              "Exit Z-Score Threshold",  f"+/- {EXIT}"),
        ("Losing Trades",      len(losses),            "Stop Loss (pts)",          STOP_LOSS),
        ("Win Rate",           f"{len(wins)/len(trades)*100:.1f}%", "Max Hold Days", MAX_HOLD),
        ("Total P&L (pts)",    round(trades.pnl_points.sum(), 2),   "Ratio Lookback (days)", RATIO_LOOKBACK),
        ("Avg P&L per Trade",  round(trades.pnl_points.mean(), 2),  "Signal Lookback (days)", LOOKBACK),
        ("Best Trade (pts)",   round(trades.pnl_points.max(), 2),   "Mode", MODE.upper()),
        ("Worst Trade (pts)",  round(trades.pnl_points.min(), 2),   "Data Source", source.upper()),
        ("Avg Holding Days",   round(trades.holding_days.mean(), 1),"", ""),
        ("", "", "", ""),
        ("EXIT REASON BREAKDOWN", "", "COUNT", "% OF TRADES"),
        ("Reverted (normal exit)", "", len(reverted), f"{len(reverted)/len(trades)*100:.0f}%"),
        ("Stop Loss hit",          "", len(stopped),  f"{len(stopped)/len(trades)*100:.0f}%"),
        ("Expiry (held max days)", "", len(expired),  f"{len(expired)/len(trades)*100:.0f}%"),
    ]

    row = 4
    for m in metrics:
        if m[0] in ("PERFORMANCE METRICS", "EXIT REASON BREAKDOWN"):
            hcell(ws, row, 1, m[0], bg=C_NAVY, sz=10)
            hcell(ws, row, 2, m[1], bg=C_NAVY, sz=10)
            hcell(ws, row, 3, m[2], bg=C_NAVY, sz=10)
            hcell(ws, row, 4, m[3], bg=C_NAVY, sz=10)
        elif m[0] == "":
            row += 1
            continue
        else:
            bg0 = C_LT_BLUE if row % 2 == 0 else C_WHITE
            dcell(ws, row, 1, m[0], bg=bg0, bold=True, align="left")
            dcell(ws, row, 2, m[1], bg=bg0)
            dcell(ws, row, 3, m[2], bg=bg0, bold=True, align="left")
            dcell(ws, row, 4, m[3], bg=bg0)
        row += 1

    for c, w in [(1,30),(2,18),(3,28),(4,18)]:
        set_col_width(ws, c, w)


# ── Sheet 2: Trade Log ────────────────────────────────────────────────────────

def write_trade_log(ws, trades):
    ws.sheet_view.showGridLines = False

    HEADERS = [
        "#",
        "Entry Date", "Exit Date", "Days Held",
        "Trade Direction",
        "What It Means",
        "SENSEX Entry", "NIFTY Entry", "Ratio Used",
        "Spread at Entry",
        "Z-Score Entry\n(Why we entered)",
        "SENSEX Exit", "NIFTY Exit",
        "Spread at Exit",
        "Z-Score Exit",
        "P&L (Points)",
        "Exit Reason",
        "RESULT",
    ]
    WIDTHS = [5, 13, 13, 10, 18, 32, 14, 14, 11, 15, 16, 14, 14, 15, 14, 14, 15, 10]

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(len(HEADERS))}1")
    t = ws["A1"]
    t.value = "TRADE LOG  —  Every Trade with Full Calculation"
    t.fill  = fill(C_NAVY)
    t.font  = font(C_WHITE, bold=True, sz=12)
    t.alignment = center()
    ws.row_dimensions[1].height = 24

    # Legend row
    leg_items = [
        (C_WIN_BG, C_WIN_FG, "WIN (Profit)"),
        (C_LOSS_BG, C_LOSS_FG, "LOSS"),
        (C_SL_BG, C_SL_FG, "Stop Loss Hit"),
        (C_EXP_BG, C_EXP_FG, "Expiry Exit"),
    ]
    ws.merge_cells("A2:B2")
    dcell(ws, 2, 1, "COLOUR LEGEND:", bg=C_LGRAY, bold=True, align="left")
    for i, (bg, fg, lbl) in enumerate(leg_items, start=3):
        dcell(ws, 2, i, lbl, bg=bg, fg=fg, bold=True)

    # Headers
    ws.row_dimensions[3].height = 40
    for ci, (h, w) in enumerate(zip(HEADERS, WIDTHS), start=1):
        hcell(ws, 3, ci, h, bg=C_BLUE)
        set_col_width(ws, ci, w)

    direction_label = {
        "LONG_SPREAD":  "BUY Sensex + SELL Nifty",
        "SHORT_SPREAD": "SELL Sensex + BUY Nifty",
    }
    direction_desc = {
        "LONG_SPREAD":  "Sensex was CHEAP vs Nifty. Buy low, wait for it to rise.",
        "SHORT_SPREAD": "Sensex was EXPENSIVE vs Nifty. Sell high, wait for it to fall.",
    }

    for idx, (_, t) in enumerate(trades.iterrows(), start=1):
        r = idx + 3
        pnl = t["pnl_points"]
        reason = t["exit_reason"]

        if reason == "stop_loss":
            bg, fg = C_SL_BG, C_SL_FG
        elif reason == "expiry":
            bg, fg = C_EXP_BG, C_EXP_FG
        elif pnl > 0:
            bg, fg = C_WIN_BG, C_WIN_FG
        else:
            bg, fg = C_LOSS_BG, C_LOSS_FG

        ws.row_dimensions[r].height = 18
        data = [
            idx,
            t["entry_date"].strftime("%d-%b-%Y") if hasattr(t["entry_date"], "strftime") else str(t["entry_date"]),
            t["exit_date"].strftime("%d-%b-%Y")   if hasattr(t["exit_date"],  "strftime") else str(t["exit_date"]),
            int(t["holding_days"]),
            direction_label.get(t["direction"], t["direction"]),
            direction_desc.get(t["direction"], ""),
            round(t["sensex_entry"], 2),
            round(t["nifty_entry"],  2),
            round(t["ratio"],        4),
            round(t["spread_entry"], 2),
            round(t["zscore_entry"], 3),
            round(t["sensex_exit"],  2),
            round(t["nifty_exit"],   2),
            round(t["spread_exit"],  2),
            round(t["zscore_exit"],  3),
            round(pnl, 2),
            reason.replace("_", " ").title(),
            "WIN  +" if pnl > 0 else "LOSS",
        ]
        for ci, val in enumerate(data, start=1):
            align = "left" if ci == 6 else "center"
            dcell(ws, r, ci, val, bg=bg, fg=fg, align=align)

    # Freeze header
    ws.freeze_panes = "A4"


# ── Sheet 3: Day-wise Log ─────────────────────────────────────────────────────

def write_daywise(ws, signals, trades):
    ws.sheet_view.showGridLines = False

    HEADERS = [
        "Date",
        "SENSEX Close", "NIFTY Close",
        "Dynamic Ratio\n(auto-calculated)",
        "Spread\n(Sensex - Ratio×Nifty)",
        "Z-Score\n(how far from normal)",
        "Market Status",
        "Trade #",
        "Direction",
        "Day P&L\n(if trade closed)",
    ]
    WIDTHS = [13, 14, 13, 16, 20, 18, 22, 10, 24, 16]

    ws.merge_cells(f"A1:{get_column_letter(len(HEADERS))}1")
    t = ws["A1"]
    t.value = "DAY-WISE LOG  —  Every Trading Day  |  Price, Ratio, Spread, Z-Score & Status"
    t.fill  = fill(C_NAVY)
    t.font  = font(C_WHITE, bold=True, sz=12)
    t.alignment = center()
    ws.row_dimensions[1].height = 24

    # Explanation row
    ws.merge_cells(f"A2:{get_column_letter(len(HEADERS))}2")
    ex = ws["A2"]
    ex.value = (
        "Z-Score > +1.5 = Sensex too expensive (SHORT signal)   |   "
        "Z-Score < -1.5 = Sensex too cheap (LONG signal)   |   "
        "Z-Score near 0 = Fair value (HOLD / EXIT)"
    )
    ex.fill = fill(C_YELLOW)
    ex.font = font(C_NAVY, bold=False, sz=9)
    ex.alignment = left()

    ws.row_dimensions[3].height = 40
    for ci, (h, w) in enumerate(zip(HEADERS, WIDTHS), start=1):
        hcell(ws, 3, ci, h, bg=C_BLUE)
        set_col_width(ws, ci, w)

    # Build trade index: date -> (trade_num, direction, pnl if exit)
    trade_open  = {}  # date -> (trade_num, direction)
    trade_exit  = {}  # date -> pnl
    for i, (_, t) in enumerate(trades.iterrows(), start=1):
        ed = pd.Timestamp(t["entry_date"])
        xd = pd.Timestamp(t["exit_date"])
        cur = ed
        while cur <= xd:
            trade_open[cur] = (i, t["direction"])
            cur += pd.Timedelta(days=1)
        trade_exit[xd] = (i, round(t["pnl_points"], 2))

    min_row = max(LOOKBACK, RATIO_LOOKBACK)

    for row_i, (_, sig) in enumerate(signals.iterrows()):
        if row_i < min_row:
            continue
        r = row_i - min_row + 4
        d  = sig["date"]
        z  = sig["zscore"]
        sp = sig["spread"]

        if pd.isna(z):
            continue

        # Status
        ts = pd.Timestamp(d)
        if ts in trade_open:
            tnum, tdir = trade_open[ts]
            status = f"IN TRADE #{tnum}"
            bg = C_LT_BLUE
        elif z >= ENTRY:
            status = "SHORT SIGNAL (Sensex rich)"
            bg = C_SL_BG
        elif z <= -ENTRY:
            status = "LONG SIGNAL  (Sensex cheap)"
            bg = C_WIN_BG
        elif z >= EXIT:
            status = "Neutral (watch short)"
            bg = C_WHITE
        elif z <= -EXIT:
            status = "Neutral (watch long)"
            bg = C_WHITE
        else:
            status = "FLAT / Hold"
            bg = C_LGRAY if row_i % 2 == 0 else C_WHITE

        day_pnl = ""
        if ts in trade_exit:
            day_pnl = trade_exit[ts][1]

        tnum_val  = trade_open[ts][0] if ts in trade_open else ""
        tdir_val  = trade_open[ts][1].replace("_", " ") if ts in trade_open else ""

        row_data = [
            d.strftime("%d-%b-%Y"),
            round(sig["sensex_close"], 2),
            round(sig["nifty_close"],  2),
            round(sig["ratio"],        4) if not pd.isna(sig["ratio"]) else "",
            round(sp, 2) if not pd.isna(sp) else "",
            round(z,  3) if not pd.isna(z)  else "",
            status,
            tnum_val,
            tdir_val,
            day_pnl,
        ]
        ws.row_dimensions[r].height = 16
        for ci, val in enumerate(row_data, start=1):
            dcell(ws, r, ci, val, bg=bg)

    ws.freeze_panes = "A4"


# ── Sheet 4: Monthly P&L ──────────────────────────────────────────────────────

def write_monthly(ws, trades):
    ws.sheet_view.showGridLines = False

    HEADERS = ["Month", "Trades", "Wins", "Losses", "Stop Loss", "Expiry",
               "Win Rate", "P&L (Points)", "Cumulative P&L"]
    WIDTHS  = [16, 10, 10, 10, 12, 10, 12, 16, 18]

    ws.merge_cells(f"A1:{get_column_letter(len(HEADERS))}1")
    t = ws["A1"]
    t.value = "MONTHLY P&L BREAKDOWN"
    t.fill  = fill(C_NAVY)
    t.font  = font(C_WHITE, bold=True, sz=12)
    t.alignment = center()
    ws.row_dimensions[1].height = 24

    ws.row_dimensions[2].height = 30
    for ci, (h, w) in enumerate(zip(HEADERS, WIDTHS), start=1):
        hcell(ws, 2, ci, h, bg=C_BLUE)
        set_col_width(ws, ci, w)

    tdf = trades.copy()
    tdf["month"] = pd.to_datetime(tdf["exit_date"]).dt.to_period("M")
    monthly = tdf.groupby("month").apply(lambda g: pd.Series({
        "trades":    len(g),
        "wins":      (g.pnl_points > 0).sum(),
        "losses":    (g.pnl_points <= 0).sum(),
        "stop_loss": (g.exit_reason == "stop_loss").sum(),
        "expiry":    (g.exit_reason == "expiry").sum(),
        "pnl":       round(g.pnl_points.sum(), 2),
    }), include_groups=False).reset_index()

    cum_pnl = 0
    for ri, (_, m) in enumerate(monthly.iterrows(), start=3):
        r   = ri
        pnl = m["pnl"]
        cum_pnl += pnl
        wr  = f"{m['wins']/m['trades']*100:.0f}%" if m["trades"] > 0 else "–"
        bg  = C_WIN_BG if pnl > 0 else (C_LOSS_BG if pnl < 0 else C_WHITE)
        fg  = C_WIN_FG if pnl > 0 else (C_LOSS_FG if pnl < 0 else None)

        row_data = [
            str(m["month"]),
            int(m["trades"]),
            int(m["wins"]),
            int(m["losses"]),
            int(m["stop_loss"]),
            int(m["expiry"]),
            wr,
            pnl,
            round(cum_pnl, 2),
        ]
        ws.row_dimensions[r].height = 18
        for ci, val in enumerate(row_data, start=1):
            dcell(ws, r, ci, val, bg=bg, fg=fg)

    # Total row
    total_r = len(monthly) + 3
    total_pnl = round(trades.pnl_points.sum(), 2)
    total_fg  = C_WIN_FG if total_pnl > 0 else C_LOSS_FG
    total_bg  = C_WIN_BG if total_pnl > 0 else C_LOSS_BG
    total_data = [
        "TOTAL",
        len(trades),
        len(trades[trades.pnl_points > 0]),
        len(trades[trades.pnl_points <= 0]),
        len(trades[trades.exit_reason == "stop_loss"]),
        len(trades[trades.exit_reason == "expiry"]),
        f"{len(trades[trades.pnl_points>0])/len(trades)*100:.0f}%",
        total_pnl, total_pnl,
    ]
    ws.row_dimensions[total_r].height = 20
    for ci, val in enumerate(total_data, start=1):
        dcell(ws, total_r, ci, val, bg=total_bg, fg=total_fg, bold=True)


# ── Main ──────────────────────────────────────────────────────────────────────

def parse_args():
    source     = "yfinance"
    start_date = None
    end_date   = None
    out_file   = f"backtest_report_{date.today()}.xlsx"
    if "--source" in sys.argv:
        idx    = sys.argv.index("--source")
        source = sys.argv[idx + 1]
    if "--start" in sys.argv:
        idx        = sys.argv.index("--start")
        start_date = sys.argv[idx + 1]
    if "--end" in sys.argv:
        idx      = sys.argv.index("--end")
        end_date = sys.argv[idx + 1]
    if "--out" in sys.argv:
        idx      = sys.argv.index("--out")
        out_file = sys.argv[idx + 1]
    return source, start_date, end_date, out_file


def main():
    source, start_date, end_date, out_file = parse_args()

    print(f"Loading data from [{source}_feed] ...")
    signals, trades = load_data(source, start_date, end_date)
    print(f"  {len(signals)} days  |  {len(trades)} trades")

    if len(trades) == 0:
        print("No trades found. Widen the date range or lower ENTRY threshold.")
        return

    summary(trades)

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Summary"
    write_summary(ws1, trades, signals, source, start_date, end_date)
    print("  Sheet 1: Summary done")

    ws2 = wb.create_sheet("Trade Log")
    write_trade_log(ws2, trades)
    print("  Sheet 2: Trade Log done")

    ws3 = wb.create_sheet("Day-wise Log")
    write_daywise(ws3, signals, trades)
    print("  Sheet 3: Day-wise Log done")

    ws4 = wb.create_sheet("Monthly P&L")
    write_monthly(ws4, trades)
    print("  Sheet 4: Monthly P&L done")

    wb.save(out_file)
    print(f"\nReport saved -> {out_file}")


if __name__ == "__main__":
    main()
